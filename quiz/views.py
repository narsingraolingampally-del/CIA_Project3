
# =========================================================
# IMPORTS
# =========================================================

import random
import string
from io import BytesIO

import pandas as pd

from django.http import (
    JsonResponse,
    HttpResponse,
)
from django.contrib.auth import get_user_model
from django.db import connection
from django.contrib import messages

from django.contrib.auth import (
    authenticate,
    login,
    logout,
    get_user_model,
)

from django.contrib.auth.decorators import (
    login_required,
    user_passes_test,
)

from django.db.models import (
    Q,
    Avg,
    Count,
)

from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)

from reportlab.lib.pagesizes import (
    A4,
    landscape,
)

from reportlab.lib.styles import (
    getSampleStyleSheet,
)

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
)

from .models import (
    User,
    Course,
    Subject,
    StudentProfile,
    FacultyProfile,
    Question,
    QuestionPaper,
    Exam,
    Result,
    ExamQuestion,
)

from .forms import (
    FacultyQuestionUploadForm,
    FacultyForm,
    StudentRegistrationForm,
    CourseForm,
    SubjectForm,
    QuestionForm,
    QuestionPaperForm,
    QuestionUploadForm,
    ExamForm,
)



# =========================================================
# ACCESS CONTROL
# =========================================================

def is_admin(user):
    return (
        user.is_authenticated
        and user.is_superuser
    )


def is_student(user):
    return (
        user.is_authenticated
        and user.is_student
    )


def is_faculty(user):
    return (
        user.is_authenticated
        and (
            user.is_faculty
            or user.is_superuser
        )
    )


def is_admin_or_faculty(user):
    return (
        user.is_authenticated
        and (
            user.is_superuser
            or user.is_staff
            or user.is_faculty
        )
    )

# =========================================
# HOME PAGE
# =========================================

def index_page(request):
    return render(request, "index.html")


# =========================================
# STUDENT LOGIN
# =========================================

def student_login(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None and user.is_student:
            login(request, user)
            return redirect("student_dashboard")

        messages.error(request, "Invalid Student Login")

    return render(request, "student/login.html")


from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect


def admin_login(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user and user.is_superuser:
            login(request, user)
            return redirect("admin_dashboard")

        messages.error(request, "Invalid Administrator Login")

    return render(request, "admin_panel/login.html")

def faculty_login(request):

    # =====================================================
    # FACULTY LOGIN
    # =====================================================

    if request.method == "POST":

        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        # =================================================
        # INVALID LOGIN
        # =================================================

        if user is None:

            messages.error(
                request,
                "Invalid username or password."
            )

            return render(
                request,
                "faculty/login.html"
            )

        # =================================================
        # ADMIN NOT ALLOWED
        # =================================================

        if user.is_superuser:

            messages.error(
                request,
                "Admin account cannot use Faculty Login. Please use Admin Login."
            )

            return render(
                request,
                "faculty/login.html"
            )

        # =================================================
        # FACULTY
        # =================================================

        if user.is_faculty:

            login(request, user)

            return redirect(
                "faculty_dashboard"
            )

        # =================================================
        # OTHER USERS
        # =================================================

        messages.error(
            request,
            "This login is only for Faculty."
        )

        return render(
            request,
            "faculty/login.html"
        )

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        "faculty/login.html"
    )
# =========================================
# LOGOUT
# =========================================

@login_required
def logout_view(request):
    logout(request)
    return redirect("index")



# =========================================
# FACULTY DASHBOARD
# =========================================
@login_required
@user_passes_test(is_faculty)
def faculty_dashboard(request):

    faculty = get_object_or_404(
        FacultyProfile,
        user=request.user
    )


    total_questions = Question.objects.count()


    total_exams = Exam.objects.count()


    total_students = StudentProfile.objects.count()


    total_results = Result.objects.count()



    recent_exams = Exam.objects.select_related(
        "course",
        "subject"
    ).order_by(
        "-id"
    )[:5]



    context = {

        "faculty": faculty,

        "total_questions": total_questions,

        "total_exams": total_exams,

        "total_students": total_students,

        "total_results": total_results,

        "recent_exams": recent_exams,

    }



    return render(

        request,

        "faculty/dashboard.html",

        context

    )


# =========================================
# TAKE QUIZ
# =========================================

# =========================================
# TAKE EXAM
# =========================================

@login_required
@user_passes_test(is_student)
def take_exam(request, exam_id):

    exam = get_object_or_404(
        Exam,
        id=exam_id,
        is_published=True
    )


    student = get_object_or_404(
        StudentProfile,
        user=request.user
    )


    # Check student belongs to exam course
    if student.course != exam.course:

        messages.error(
            request,
            "You are not allowed to attend this exam."
        )

        return redirect(
            "student_dashboard"
        )


    exam_questions = exam.exam_questions.select_related(
        "question"
    ).all()


    return render(
        request,
        "student/take_exam.html",
        {
            "exam": exam,
            "exam_questions": exam_questions,
            "duration_seconds": exam.duration * 60,
        },
    )
# =========================================
# SUBMIT QUIZ
# =========================================

# =========================================
# SUBMIT EXAM
# =========================================


    
# =========================================
# VIEW QUESTIONS
# =========================================

# =========================================================
# FACULTY - VIEW QUESTIONS
# =========================================================

@login_required
@user_passes_test(is_faculty)
def view_questions(request):

    # =====================================================
    # GET FILTER VALUES
    # =====================================================

    course_id = request.GET.get("course", "").strip()

    semester = request.GET.get("semester", "").strip()

    subject_id = request.GET.get("subject", "").strip()

    search = request.GET.get("search", "").strip()

    # =====================================================
    # GET ALL QUESTIONS
    # =====================================================

    questions = Question.objects.select_related(
        "course",
        "subject",
    ).all().order_by(
        "course__name",
        "semester",
        "subject__code",
        "id",
    )

    # =====================================================
    # COURSE FILTER
    # =====================================================

    if course_id:

        questions = questions.filter(
            course_id=course_id
        )

    # =====================================================
    # SEMESTER FILTER
    # =====================================================

    if semester:

        questions = questions.filter(
            semester=semester
        )

    # =====================================================
    # SUBJECT FILTER
    # =====================================================

    if subject_id:

        questions = questions.filter(
            subject_id=subject_id
        )

    # =====================================================
    # SEARCH FILTER
    # =====================================================

    if search:

        questions = questions.filter(

            Q(
                question_text__icontains=search
            )

            |

            Q(
                option1__icontains=search
            )

            |

            Q(
                option2__icontains=search
            )

            |

            Q(
                option3__icontains=search
            )

            |

            Q(
                option4__icontains=search
            )

            |

            Q(
                subject__name__icontains=search
            )

            |

            Q(
                subject__code__icontains=search
            )

        )

    # =====================================================
    # COURSES
    # =====================================================

    courses = Course.objects.all().order_by(
        "name"
    )

    # =====================================================
    # SUBJECTS
    #
    # Initially show all subjects.
    # JavaScript/filtering can narrow them later.
    # =====================================================

    subjects = Subject.objects.all().order_by(
        "code",
        "name"
    )

    # =====================================================
    # TOTAL QUESTIONS
    # =====================================================

    total_questions = questions.count()

    # =====================================================
    # DEBUG
    # =====================================================

    print("=" * 60)
    print("FACULTY QUESTION BANK")
    print("=" * 60)

    print("Course filter:", course_id)
    print("Semester filter:", semester)
    print("Subject filter:", subject_id)
    print("Search:", search)

    print(
        "Questions found:",
        total_questions
    )

    print("=" * 60)

    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        "faculty/view_questions.html",
        {
            "questions": questions,
            "courses": courses,
            "subjects": subjects,
            "total_questions": total_questions,

            # Keep filter values selected
            "selected_course": course_id,
            "selected_semester": semester,
            "selected_subject": subject_id,
            "search": search,
        }
    )
# =========================================
# VIEW RESULTS
# =========================================
@login_required
@user_passes_test(is_faculty)
def view_results(request):

    results = Result.objects.select_related(
        "student",
        "student__studentprofile",
        "exam",
        "exam__subject",
        "exam__course"
    ).order_by("-completed_at")


    context = {

        "results": results

    }


    return render(
        request,
        "faculty/view_results.html",
        context
    )
# =========================================
# ADMIN DASHBOARD
# =========================================

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):

    student_count = StudentProfile.objects.count()
    faculty_count = FacultyProfile.objects.count()
    course_count = Course.objects.count()
    subject_count = Subject.objects.count()
    question_count = Question.objects.count()
    exam_count = Exam.objects.count()
    published_exam_count = Exam.objects.filter(
        is_published=True
    ).count()
    result_count = Result.objects.count()

    recent_exams = Exam.objects.select_related(
        "course",
        "subject"
    ).order_by("-id")[:5]

    # Students by Course
    courses = Course.objects.all()

    course_names = []
    course_students = []

    for course in courses:

        course_names.append(course.name)

        course_students.append(
            StudentProfile.objects.filter(
                course=course
            ).count()
        )

    # Pass / Fail
    pass_count = Result.objects.filter(
        percentage__gte=50
    ).count()

    fail_count = Result.objects.filter(
        percentage__lt=50
    ).count()

    context = {

        "student_count": student_count,
        "faculty_count": faculty_count,
        "course_count": course_count,
        "subject_count": subject_count,
        "question_count": question_count,
        "exam_count": exam_count,
        "published_exam_count": published_exam_count,
        "result_count": result_count,
        "recent_exams": recent_exams,

        "course_names": course_names,
        "course_students": course_students,

        "pass_count": pass_count,
        "fail_count": fail_count,

    }

    return render(
        request,
        "admin_panel/dashboard.html",
        context
    )
# ==================================================
# ADD QUESTION
# ==================================================

@login_required
@user_passes_test(is_admin)
def add_question(request):

    if request.method == "POST":

        form = QuestionForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Question added successfully."
            )

            return redirect("question_bank")


    else:

        form = QuestionForm()


    return render(
        request,
        "admin_panel/add_question.html",
        {
            "form": form
        }
    )


# =========================================
# FACULTY MANAGEMENT
# =========================================

from django.db.models import Q

@login_required
@user_passes_test(is_admin)
def manage_faculty(request):

    search = request.GET.get("search", "")

    faculty = FacultyProfile.objects.select_related(
        "user",
        "course"
    ).prefetch_related(
        "subjects"
    )

    if search:
        faculty = faculty.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )

    faculty = faculty.order_by("user__username")

    return render(
        request,
        "admin_panel/faculty_list.html",
        {
            "faculty": faculty,
            "search": search,
        },
    )


# =========================================
# ADD FACULTY
# =========================================



@login_required
@user_passes_test(is_admin)
def add_faculty(request):

    if request.method == "POST":

        form = FacultyForm(request.POST)

        if form.is_valid():

            username = form.cleaned_data["username"].strip()

            # ------------------------------------------
            # DUPLICATE USERNAME
            # ------------------------------------------

            if User.objects.filter(
                username=username
            ).exists():

                form.add_error(
                    "username",
                    "This username already exists."
                )

                return render(
                    request,
                    "admin_panel/add_faculty.html",
                    {
                        "form": form
                    }
                )

            # ------------------------------------------
            # CREATE USER
            # ------------------------------------------

            try:

                user = User.objects.create_user(
                    username=username,
                    email=form.cleaned_data["email"].strip(),
                    password=form.cleaned_data["password"],
                    first_name=form.cleaned_data["first_name"].strip(),
                    last_name=form.cleaned_data["last_name"].strip(),
                    is_faculty=True
                )

                # --------------------------------------
                # CREATE FACULTY PROFILE
                # --------------------------------------

                FacultyProfile.objects.create(
                    user=user,
                    department=form.cleaned_data["department"].strip()
                )

                messages.success(
                    request,
                    f"Faculty '{username}' added successfully."
                )

                return redirect(
                    "manage_faculty"
                )

            except Exception as e:

                messages.error(
                    request,
                    f"Unable to add faculty: {e}"
                )

                return render(
                    request,
                    "admin_panel/add_faculty.html",
                    {
                        "form": form
                    }
                )

    else:

        form = FacultyForm()

    return render(
        request,
        "admin_panel/add_faculty.html",
        {
            "form": form
        }
    )# =========================================
# EDIT FACULTY
# =========================================

@login_required
@user_passes_test(is_admin)
def edit_faculty(request, id):

    faculty = get_object_or_404(
        FacultyProfile,
        id=id
    )

    user = faculty.user

    if request.method == "POST":

        username = request.POST.get(
            "username",
            ""
        ).strip()

        first_name = request.POST.get(
            "first_name",
            ""
        ).strip()

        last_name = request.POST.get(
            "last_name",
            ""
        ).strip()

        email = request.POST.get(
            "email",
            ""
        ).strip()

        department = request.POST.get(
            "department",
            ""
        ).strip()

        new_password = request.POST.get(
            "new_password",
            ""
        )

        confirm_password = request.POST.get(
            "confirm_password",
            ""
        )

        # ------------------------------------------
        # USERNAME
        # ------------------------------------------

        if not username:

            messages.error(
                request,
                "Username is required."
            )

            return redirect(
                "edit_faculty",
                id=id
            )

        if User.objects.filter(
            username=username
        ).exclude(
            id=user.id
        ).exists():

            messages.error(
                request,
                "This username already exists."
            )

            return redirect(
                "edit_faculty",
                id=id
            )

        # ------------------------------------------
        # PASSWORD
        # ------------------------------------------

        if new_password:

            if len(new_password) < 8:

                messages.error(
                    request,
                    "Password must be at least 8 characters."
                )

                return redirect(
                    "edit_faculty",
                    id=id
                )

            if new_password != confirm_password:

                messages.error(
                    request,
                    "Passwords do not match."
                )

                return redirect(
                    "edit_faculty",
                    id=id
                )

            user.set_password(
                new_password
            )

        # ------------------------------------------
        # UPDATE USER
        # ------------------------------------------

        user.username = username
        user.first_name = first_name
        user.last_name = last_name
        user.email = email

        user.save()

        # ------------------------------------------
        # UPDATE FACULTY PROFILE
        # ------------------------------------------

        faculty.department = department

        faculty.save()

        messages.success(
            request,
            "Faculty updated successfully."
        )

        return redirect(
            "manage_faculty"
        )

    return render(
        request,
        "admin_panel/edit_faculty.html",
        {
            "faculty": faculty,
            "user": user
        }
    )
# DELETE FACULTY
# =========================================

@login_required
@user_passes_test(is_admin)
def delete_faculty(request, id):

    faculty = get_object_or_404(
        FacultyProfile,
        id=id
    )

    if request.method == "POST":

        username = faculty.user.username

        faculty.user.delete()

        messages.success(
            request,
            f"Faculty '{username}' deleted successfully."
        )

        return redirect(
            "manage_faculty"
        )

    return render(
        request,
        "admin_panel/delete_faculty.html",
        {
            "faculty": faculty
        }
    )
# =========================================
# CREATE EXAM
# =========================================
# =========================================
# FACULTY CREATE EXAM
# =========================================

@login_required
@user_passes_test(is_faculty)
def create_exam(request):

    if request.method == "POST":

        form = ExamForm(request.POST)

        if form.is_valid():

            exam = form.save(commit=False)

            # Faculty-created exams should initially be unpublished
            exam.is_published = False
            exam.save()

            print("==============================")
            print("EXAM CREATED")
            print("ID:", exam.id)
            print("COURSE:", exam.course)
            print("COURSE ID:", exam.course_id)
            print("SEMESTER:", exam.semester)
            print("SUBJECT:", exam.subject)
            print("SUBJECT ID:", exam.subject_id)
            print("NUMBER OF QUESTIONS:", exam.number_of_questions)
            print("==============================")


            # ==========================================
            # GET QUESTIONS
            # COURSE + SEMESTER + SUBJECT
            # ==========================================

            questions = Question.objects.filter(
                course_id=exam.course_id,
                semester=exam.semester,
                subject_id=exam.subject_id
            )

            available_questions = questions.count()

            print(
                "AVAILABLE QUESTIONS:",
                available_questions
            )


            # ==========================================
            # CHECK QUESTION COUNT
            # ==========================================

            if available_questions < exam.number_of_questions:

                messages.error(
                    request,
                    f"Only {available_questions} questions available "
                    f"for {exam.subject} in Semester {exam.semester}. "
                    f"You requested {exam.number_of_questions}."
                )

                exam.delete()

                return redirect("create_exam")


            # ==========================================
            # RANDOM SELECTION
            # ==========================================

            selected_questions = random.sample(
                list(questions),
                exam.number_of_questions
            )


            # ==========================================
            # CREATE EXAM QUESTIONS
            # ==========================================

            for question in selected_questions:

                ExamQuestion.objects.create(
                    exam=exam,
                    question=question
                )


            # ==========================================
            # SUCCESS
            # ==========================================

            messages.success(
                request,
                f"Exam '{exam.exam_name}' created successfully "
                f"with {exam.number_of_questions} questions."
            )

            return redirect("faculty_dashboard")

        else:

            print("EXAM FORM ERRORS:")
            print(form.errors)

    else:

        form = ExamForm()


    return render(
        request,
        "faculty/create_exam.html",
        {
            "form": form
        }
    )# =========================================
# ADD STUDENT
# =========================================

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render

from .models import (
    Exam,
    ExamQuestion,
    Result,
    StudentAnswer,
)
@login_required
@user_passes_test(is_student)
def submit_exam(request):

    # =====================================================
    # ONLY POST REQUEST
    # =====================================================

    if request.method != "POST":

        return redirect(
            "student_dashboard"
        )


    # =====================================================
    # GET EXAM ID
    # =====================================================

    exam_id = request.POST.get(
        "exam_id"
    )


    if not exam_id:

        messages.error(
            request,
            "Invalid exam submission."
        )

        return redirect(
            "student_dashboard"
        )


    # =====================================================
    # GET SUBMISSION REASON
    # =====================================================

    submission_reason = request.POST.get(
        "submission_reason",
        "manual"
    )


    # =====================================================
    # ALLOWED SUBMISSION REASONS
    # =====================================================

    allowed_reasons = {
        "manual",
        "tab_switch",
        "window_switch",
        "browser_back",
        "time_expired",
    }


    if submission_reason not in allowed_reasons:

        submission_reason = "manual"


    # =====================================================
    # GET EXAM
    # =====================================================

    exam = get_object_or_404(
        Exam,
        id=exam_id
    )


    # =====================================================
    # PREVENT MULTIPLE ATTEMPTS
    # =====================================================

    if Result.objects.filter(
        student=request.user,
        exam=exam
    ).exists():

        messages.warning(
            request,
            "You have already attempted this exam."
        )

        return redirect(
            "student_dashboard"
        )


    # =====================================================
    # GET QUESTIONS
    # =====================================================

    exam_questions = (
        ExamQuestion.objects
        .filter(exam=exam)
        .select_related("question")
        .order_by("id")
    )


    # =====================================================
    # VARIABLES
    # =====================================================

    score = 0

    total_marks = 0

    student_answers = []


    # =====================================================
    # PROCESS EACH QUESTION
    # =====================================================

    for eq in exam_questions:

        question = eq.question


        # -------------------------------------------------
        # MARKS
        # -------------------------------------------------

        question_marks = (
            question.marks
            if question.marks
            else 1
        )


        total_marks += question_marks


        # -------------------------------------------------
        # GET SELECTED ANSWER
        # -------------------------------------------------

        selected = request.POST.get(
            f"question_{question.id}"
        )


        if selected:

            selected = str(
                selected
            ).strip()

        else:

            selected = ""


        # -------------------------------------------------
        # CORRECT ANSWER
        # -------------------------------------------------

        correct_answer = str(
            question.correct_answer
        ).strip()


        # -------------------------------------------------
        # CHECK ANSWER
        # -------------------------------------------------

        is_correct = (
            selected != ""
            and
            selected == correct_answer
        )


        # -------------------------------------------------
        # MARKS OBTAINED
        # -------------------------------------------------

        if is_correct:

            marks_obtained = (
                question_marks
            )

            score += question_marks

        else:

            marks_obtained = 0


        # =================================================
        # DEBUG
        # =================================================

        print("----------------------------------------")

        print(
            "Question ID:",
            question.id
        )

        print(
            "ExamQuestion ID:",
            eq.id
        )

        print(
            "Selected:",
            selected
        )

        print(
            "Correct:",
            correct_answer
        )

        print(
            "Is Correct:",
            is_correct
        )

        print(
            "Marks:",
            marks_obtained
        )


        # =================================================
        # SAVE STUDENT ANSWER
        # =================================================

        student_answers.append(

            StudentAnswer(

                student=request.user,

                exam=exam,

                question=question,

                selected_answer=selected,

                is_correct=is_correct,

                marks_obtained=marks_obtained

            )

        )


    # =====================================================
    # PERCENTAGE
    # =====================================================

    if total_marks > 0:

        percentage = (
            score /
            total_marks
        ) * 100

    else:

        percentage = 0


    # =====================================================
    # SAVE RESULT + ANSWERS
    # =====================================================

    with transaction.atomic():

        result = Result.objects.create(

            student=request.user,

            exam=exam,

            score=score,

            total_marks=total_marks,

            percentage=percentage

        )


        StudentAnswer.objects.bulk_create(
            student_answers
        )


    # =====================================================
    # FINAL DEBUG
    # =====================================================

    print("========================================")

    print(
        "STUDENT:",
        request.user.username
    )

    print(
        "EXAM:",
        exam.exam_name
    )

    print(
        "EXAM ID:",
        exam.id
    )

    print(
        "SUBMISSION REASON:",
        submission_reason
    )

    print(
        "SCORE:",
        score
    )

    print(
        "TOTAL MARKS:",
        total_marks
    )

    print(
        "PERCENTAGE:",
        percentage
    )

    print(
        "ANSWERS SAVED:",
        len(student_answers)
    )

    print("========================================")


    # =====================================================
    # RESULT PAGE
    # =====================================================

    return render(
        request,
        "student/result.html",
        {
            "exam": exam,

            "result": result,

            "score": score,

            "total_marks": total_marks,

            "percentage": round(
                percentage,
                2
            ),

            "submission_reason":
                submission_reason,
        }
    )


@login_required
@user_passes_test(is_admin)
def delete_student(request, id):

    student = get_object_or_404(
        StudentProfile,
        id=id
    )

    student.user.delete()

    messages.success(
        request,
        "Student deleted successfully."
    )

    return redirect("manage_students")


# =========================================
# UPLOAD STUDENTS
# =======================================




# ==========================================================
# BULK UPLOAD STUDENTS
# ==========================================================

@login_required
@user_passes_test(is_admin)
def upload_students(request):

    courses = Course.objects.all()

    if request.method == "POST":

        excel_file = request.FILES.get("excel_file")
        course_id = request.POST.get("course")
        semester = request.POST.get("semester")
        academic_year = request.POST.get("academic_year")

        # -----------------------------------------
        # BASIC VALIDATION
        # -----------------------------------------

        if not excel_file:
            messages.error(
                request,
                "Please select an Excel file."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        if not course_id:
            messages.error(
                request,
                "Please select a course."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        if not semester:
            messages.error(
                request,
                "Please select a semester."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        if not academic_year:
            messages.error(
                request,
                "Please enter the academic year."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        # -----------------------------------------
        # COURSE
        # -----------------------------------------

        try:

            course = Course.objects.get(
                id=course_id
            )

        except Course.DoesNotExist:

            messages.error(
                request,
                "Selected course does not exist."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        # -----------------------------------------
        # SEMESTER VALIDATION
        # -----------------------------------------

        try:

            semester = int(semester)

        except ValueError:

            messages.error(
                request,
                "Invalid semester."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        if semester < 1 or semester > 8:

            messages.error(
                request,
                "Semester must be between 1 and 8."
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        # -----------------------------------------
        # READ EXCEL
        # -----------------------------------------

        try:

            df = pd.read_excel(
                excel_file,
                dtype=str
            )

        except Exception as e:

            messages.error(
                request,
                f"Unable to read Excel file: {e}"
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        # -----------------------------------------
        # CLEAN COLUMN NAMES
        # -----------------------------------------

        df.columns = (
            df.columns
            .astype(str)
            .str.strip()
            .str.lower()
        )

        # -----------------------------------------
        # REQUIRED COLUMNS
        # -----------------------------------------

        required_columns = [
            "username",
            "name",
            "aadhaar number",
        ]

        missing_columns = [
            column
            for column in required_columns
            if column not in df.columns
        ]

        if missing_columns:

            messages.error(
                request,
                "Missing Excel columns: "
                + ", ".join(missing_columns)
            )

            return render(
                request,
                "admin_panel/upload_students.html",
                {
                    "courses": courses
                }
            )

        # -----------------------------------------
        # REMOVE EMPTY ROWS
        # -----------------------------------------

        df = df.dropna(
            how="all"
        )

        # -----------------------------------------
        # COUNTERS
        # -----------------------------------------

        created_count = 0
        skipped_count = 0

        errors = []

        # -----------------------------------------
        # PROCESS STUDENTS
        # -----------------------------------------

        for index, row in df.iterrows():

            excel_row = index + 2

            username = str(
                row["username"]
            ).strip()

            name = str(
                row["name"]
            ).strip()

            aadhaar = str(
                row["aadhaar number"]
            ).strip()

            # -------------------------------------
            # EMPTY VALUES
            # -------------------------------------

            if not username or username.lower() == "nan":

                errors.append(
                    f"Row {excel_row}: Username is required."
                )

                skipped_count += 1

                continue

            if not name or name.lower() == "nan":

                errors.append(
                    f"Row {excel_row}: Name is required."
                )

                skipped_count += 1

                continue

            if not aadhaar or aadhaar.lower() == "nan":

                errors.append(
                    f"Row {excel_row}: Aadhaar number is required."
                )

                skipped_count += 1

                continue

            # -------------------------------------
            # AADHAAR VALIDATION
            # -------------------------------------

            aadhaar = aadhaar.replace(
                " ",
                ""
            )

            if not aadhaar.isdigit():

                errors.append(
                    f"Row {excel_row}: "
                    f"Aadhaar must contain only digits."
                )

                skipped_count += 1

                continue

            if len(aadhaar) != 12:

                errors.append(
                    f"Row {excel_row}: "
                    f"Aadhaar must contain exactly 12 digits."
                )

                skipped_count += 1

                continue

            # -------------------------------------
            # USERNAME DUPLICATE CHECK
            # -------------------------------------

            if User.objects.filter(
                username=username
            ).exists():

                errors.append(
                    f"Row {excel_row}: "
                    f"Username '{username}' already exists."
                )

                skipped_count += 1

                continue

            # -------------------------------------
            # AADHAAR DUPLICATE CHECK
            # -------------------------------------

            if StudentProfile.objects.filter(
                aadhaar_number=aadhaar
            ).exists():

                errors.append(
                    f"Row {excel_row}: "
                    f"Aadhaar number already exists."
                )

                skipped_count += 1

                continue

            # -------------------------------------
            # PASSWORD
            # -------------------------------------

            password = f"CIA@{username}"

            # -------------------------------------
            # CREATE USER
            # -------------------------------------

            try:

                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=name,
                    is_student=True
                )

                # ---------------------------------
                # CREATE STUDENT PROFILE
                # ---------------------------------

                StudentProfile.objects.create(
                    user=user,
                    course=course,
                    semester=semester,
                    academic_year=academic_year,
                    aadhaar_number=aadhaar
                )

                created_count += 1

            except Exception as e:

                errors.append(
                    f"Row {excel_row}: {str(e)}"
                )

                # Remove partially-created user
                try:

                    if user:
                        user.delete()

                except Exception:
                    pass

                skipped_count += 1

        # -----------------------------------------
        # RESULT MESSAGE
        # -----------------------------------------

        if created_count > 0:

            messages.success(
                request,
                f"{created_count} student(s) uploaded successfully."
            )

        if skipped_count > 0:

            messages.warning(
                request,
                f"{skipped_count} student(s) were skipped."
            )

        # -----------------------------------------
        # SHOW ERRORS
        # -----------------------------------------

        for error in errors:

            messages.error(
                request,
                error
            )

        return render(
            request,
            "admin_panel/upload_students.html",
            {
                "courses": courses
            }
        )

    # ---------------------------------------------
    # GET REQUEST
    # ---------------------------------------------

    return render(
        request,
        "admin_panel/upload_students.html",
        {
            "courses": courses
        }
    )

# ==========================================================
# DOWNLOAD STUDENT EXCEL TEMPLATE
# ==========================================================

def download_student_template(request):

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"

    # Headers
    headers = [
        "Username",
        "Name",
        "Aadhaar Number",
    ]

    worksheet.append(headers)

    # Example rows
    worksheet.append([
        "107225861001",
        "Student Name",
        "123456789012",
    ])

    worksheet.append([
        "107225861002",
        "Student Name 2",
        "234567890123",
    ])

    # Header formatting
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    # Column widths
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 20

    # Keep Aadhaar as text
    for row in worksheet.iter_rows(
        min_row=2,
        min_col=3,
        max_col=3
    ):
        for cell in row:
            cell.number_format = "@"

    # Response
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_upload_template.xlsx"'
    )

    workbook.save(response)

    return response# =========================================
# COURSE MANAGEMENT
# =========================================

@login_required
@user_passes_test(is_admin)
def manage_courses(request):

    courses = Course.objects.all().order_by("id")

    context = {
        "courses": courses,
        "total_courses": courses.count(),
    }

    return render(
        request,
        "admin_panel/manage_courses.html",
        context
    )


@login_required
@user_passes_test(is_admin)
def add_course(request):

    if request.method == "POST":

        form = CourseForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Course added successfully."
            )

            return redirect("manage_courses")

    else:

        form = CourseForm()

    return render(
        request,
        "admin_panel/add_course.html",
        {
            "form": form
        }
    )


@login_required
@user_passes_test(is_admin)
def edit_course(request, id):

    course = get_object_or_404(
        Course,
        id=id
    )

    if request.method == "POST":

        form = CourseForm(
            request.POST,
            instance=course
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Course updated successfully."
            )

            return redirect("manage_courses")

    else:

        form = CourseForm(instance=course)

    return render(
        request,
        "admin_panel/add_course.html",
        {
            "form": form
        }
    )


@login_required
@user_passes_test(is_admin)
def delete_course(request, id):

    course = get_object_or_404(
        Course,
        id=id
    )

    course.delete()

    messages.success(
        request,
        "Course deleted successfully."
    )

    return redirect("manage_courses")


# =========================================
# SUBJECT MANAGEMENT
# =========================================

@login_required
@user_passes_test(is_admin)
def manage_subjects(request):

    subjects = Subject.objects.select_related(
        "course"
    ).all().order_by("id")

    context = {

        "subjects": subjects,

        "total_subjects": subjects.count(),

    }

    return render(
        request,
        "admin_panel/manage_subjects.html",
        context
    )


@login_required
@user_passes_test(is_admin)
def add_subject(request):

    if request.method == "POST":

        form = SubjectForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Subject added successfully."
            )

            return redirect("manage_subjects")

    else:

        form = SubjectForm()

    return render(
        request,
        "admin_panel/add_subject.html",
        {
            "form": form
        }
    )


@login_required
@user_passes_test(is_admin)
def edit_subject(request, id):

    subject = get_object_or_404(
        Subject,
        id=id
    )

    if request.method == "POST":

        form = SubjectForm(
            request.POST,
            instance=subject
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Subject updated successfully."
            )

            return redirect("manage_subjects")

    else:

        form = SubjectForm(instance=subject)

    return render(
        request,
        "admin_panel/add_subject.html",
        {
            "form": form
        }
    )


@login_required
@user_passes_test(is_admin)
def delete_subject(request, id):

    subject = get_object_or_404(
        Subject,
        id=id
    )

    subject.delete()

    messages.success(
        request,
        "Subject deleted successfully."
    )

    return redirect("manage_subjects")
# =========================================
# QUESTION BANK
# =========================================
@login_required
@user_passes_test(is_admin)
def question_bank(request):

    # =====================================================
    # ALL QUESTIONS
    # =====================================================

    questions = (
        Question.objects
        .select_related(
            "course",
            "subject"
        )
        .all()
        .order_by("-id")
    )


    # =====================================================
    # DROPDOWN DATA
    # =====================================================

    courses = Course.objects.all().order_by("name")

    subjects = (
        Subject.objects
        .select_related("course")
        .all()
        .order_by("code", "name")
    )


    # =====================================================
    # SEMESTER CHOICES
    # =====================================================

    semester_choices = [
        ("1", "Semester 1"),
        ("2", "Semester 2"),
        ("3", "Semester 3"),
        ("4", "Semester 4"),
        ("5", "Semester 5"),
        ("6", "Semester 6"),
        ("7", "Semester 7"),
        ("8", "Semester 8"),
    ]


    # =====================================================
    # GET FILTER VALUES
    # =====================================================

    search = request.GET.get("search", "").strip()

    selected_course = request.GET.get(
        "course",
        ""
    ).strip()

    selected_subject = request.GET.get(
        "subject",
        ""
    ).strip()

    selected_semester = request.GET.get(
        "semester",
        ""
    ).strip()


    # =====================================================
    # SEARCH FILTER
    # =====================================================

    if search:

        questions = questions.filter(
            question_text__icontains=search
        )


    # =====================================================
    # COURSE FILTER
    # =====================================================

    if selected_course:

        questions = questions.filter(
            course_id=selected_course
        )


    # =====================================================
    # SUBJECT FILTER
    # =====================================================

    if selected_subject:

        questions = questions.filter(
            subject_id=selected_subject
        )


    # =====================================================
    # SEMESTER FILTER
    # =====================================================

    if selected_semester:

        questions = questions.filter(
            semester=selected_semester
        )


    # =====================================================
    # CONTEXT
    # =====================================================

    context = {

        "questions": questions,

        "courses": courses,

        "subjects": subjects,

        "semester_choices": semester_choices,

        "search": search,

        "selected_course": selected_course,

        "selected_subject": selected_subject,

        "selected_semester": selected_semester,

    }


    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        "admin_panel/question_bank.html",
        context
    )

# =========================================
# UPLOAD QUESTIONS
# =========================================

@login_required
@user_passes_test(is_admin_or_faculty)
def upload_questions(request):

    # =====================================================
    # GET
    # =====================================================

    if request.method == "GET":

        form = QuestionUploadForm()

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # POST
    # =====================================================

    form = QuestionUploadForm(
        request.POST,
        request.FILES
    )

    print("\n========================================")
    print("QUESTION BANK UPLOAD")
    print("========================================")

    print("POST DATA:")
    print(request.POST)

    print("\nFILES:")
    print(request.FILES)

    # =====================================================
    # FORM VALIDATION
    # =====================================================

    if not form.is_valid():

        print("\n========================================")
        print("FORM INVALID")
        print("========================================")

        print(form.errors.as_json())

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    print("\nFORM IS VALID")

    # =====================================================
    # FORM DATA
    # =====================================================

    course = form.cleaned_data["course"]

    semester = int(
        form.cleaned_data["semester"]
    )

    subject = form.cleaned_data["subject"]

    academic_year = (
        form.cleaned_data["academic_year"]
        .strip()
    )

    excel_file = form.cleaned_data["excel_file"]

    print("\n----------------------------------------")
    print("SELECTED DATA")
    print("----------------------------------------")

    print("Course:", course)
    print("Course ID:", course.id)

    print("Semester:", semester)

    print("Subject:", subject)
    print("Subject ID:", subject.id)

    print("Academic Year:", academic_year)

    print("Excel:", excel_file.name)

    # =====================================================
    # COURSE / SUBJECT VALIDATION
    # =====================================================

    if subject.course_id != course.id:

        form.add_error(
            "subject",
            "Selected subject does not belong to "
            "the selected course."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # SEMESTER / SUBJECT VALIDATION
    # =====================================================

    if int(subject.semester) != semester:

        form.add_error(
            "subject",
            "Selected subject does not belong to "
            "the selected semester."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # CHECK FILE EXTENSION
    # =====================================================

    filename = excel_file.name.lower()

    if not filename.endswith(
        (".xlsx", ".xls")
    ):

        messages.error(
            request,
            "Please upload a valid Excel file (.xlsx or .xls)."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # READ EXCEL
    # =====================================================

    try:

        df = pd.read_excel(
            excel_file
        )

    except Exception as e:

        print("\nEXCEL READ ERROR:", e)

        messages.error(
            request,
            f"Unable to read Excel file: {e}"
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # CLEAN COLUMN NAMES
    # =====================================================

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
    )

    print("\n========================================")
    print("EXCEL INFORMATION")
    print("========================================")

    print("Columns:")
    print(list(df.columns))

    print("Number of rows:", len(df))

    # =====================================================
    # REQUIRED COLUMNS
    # =====================================================

    required_columns = [

        "question_text",

        "option1",

        "option2",

        "option3",

        "option4",

        "correct_answer",

        "marks",

    ]

    missing_columns = [

        column

        for column in required_columns

        if column not in df.columns

    ]

    if missing_columns:

        messages.error(
            request,
            "Missing Excel columns: "
            + ", ".join(missing_columns)
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # KEEP ONLY REQUIRED COLUMNS
    # =====================================================

    df = df[
        required_columns
    ]

    # =====================================================
    # REMOVE COMPLETELY EMPTY ROWS
    # =====================================================

    df = df.dropna(
        how="all"
    )

    if df.empty:

        messages.error(
            request,
            "The Excel file does not contain "
            "any questions."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # VALIDATE ALL ROWS FIRST
    #
    # IMPORTANT:
    # No database records are created here.
    # =====================================================

    valid_questions = []

    errors = []

    print("\n========================================")
    print("VALIDATING QUESTIONS")
    print("========================================")

    for row_number, row in df.iterrows():

        # Excel header is row 1
        # Therefore first data row is row 2

        excel_row = row_number + 2

        # =================================================
        # QUESTION TEXT
        # =================================================

        question_value = row[
            "question_text"
        ]

        if pd.isna(question_value):

            errors.append(
                f"Excel row {excel_row}: "
                "question_text is empty."
            )

            continue

        question_text = str(
            question_value
        ).strip()

        if not question_text:

            errors.append(
                f"Excel row {excel_row}: "
                "question_text is empty."
            )

            continue

        # =================================================
        # OPTIONS
        # =================================================

        option_values = {}

        option_error = False

        for column in [

            "option1",
            "option2",
            "option3",
            "option4",

        ]:

            value = row[column]

            if pd.isna(value):

                value = ""

            else:

                value = str(
                    value
                ).strip()

            option_values[column] = value

            if not value:

                errors.append(
                    f"Excel row {excel_row}: "
                    f"{column} is empty."
                )

                option_error = True

        if option_error:

            continue

        # =================================================
        # CORRECT ANSWER
        # =================================================

        correct_value = row[
            "correct_answer"
        ]

        if pd.isna(correct_value):

            errors.append(
                f"Excel row {excel_row}: "
                "correct_answer is empty."
            )

            continue

        try:

            correct_answer = str(
                int(
                    float(
                        correct_value
                    )
                )
            )

        except (
            ValueError,
            TypeError
        ):

            correct_answer = str(
                correct_value
            ).strip()

        # =================================================
        # VALID CORRECT ANSWER
        # =================================================

        if correct_answer not in [

            "1",
            "2",
            "3",
            "4",

        ]:

            errors.append(
                f"Excel row {excel_row}: "
                f"Invalid correct_answer "
                f"'{correct_answer}'. "
                f"Use 1, 2, 3 or 4."
            )

            continue

        # =================================================
        # MARKS
        # =================================================

        marks_value = row[
            "marks"
        ]

        if pd.isna(marks_value):

            errors.append(
                f"Excel row {excel_row}: "
                "marks is empty."
            )

            continue

        try:

            marks_float = float(
                marks_value
            )

            marks = int(
                marks_float
            )

            # Prevent 1.5 from silently becoming 1

            if marks_float != marks:

                raise ValueError

        except (
            ValueError,
            TypeError
        ):

            errors.append(
                f"Excel row {excel_row}: "
                f"Invalid marks '{marks_value}'. "
                f"Marks must be a whole number."
            )

            continue

        # =================================================
        # MARKS MUST BE GREATER THAN ZERO
        # =================================================

        if marks <= 0:

            errors.append(
                f"Excel row {excel_row}: "
                "marks must be greater than 0."
            )

            continue

        # =================================================
        # PREPARE QUESTION
        #
        # Do NOT save yet.
        # =================================================

        valid_questions.append(

            Question(

                course=course,

                subject=subject,

                academic_year=academic_year,

                semester=semester,

                uploaded_by=request.user,

                question_text=question_text,

                option1=option_values["option1"],

                option2=option_values["option2"],

                option3=option_values["option3"],

                option4=option_values["option4"],

                correct_answer=correct_answer,

                marks=marks,

            )

        )

    # =====================================================
    # VALIDATION ERRORS
    # =====================================================

    if errors:

        print("\n========================================")
        print("EXCEL VALIDATION ERRORS")
        print("========================================")

        for error in errors:

            print(error)

        # -------------------------------------------------
        # Show first several errors
        # -------------------------------------------------

        for error in errors[:10]:

            messages.error(
                request,
                error
            )

        if len(errors) > 10:

            messages.error(
                request,
                f"And {len(errors) - 10} "
                f"more error(s) found."
            )

        messages.error(
            request,
            "Upload cancelled. "
            "No questions were added to the database."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # NO VALID QUESTIONS
    # =====================================================

    if not valid_questions:

        messages.error(
            request,
            "No valid questions were found "
            "in the Excel file."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # DUPLICATE CHECK
    #
    # Prevent uploading the exact same question twice.
    # =====================================================

    duplicate_questions = []

    for question in valid_questions:

        exists = Question.objects.filter(

            course=course,

            subject=subject,

            semester=semester,

            academic_year=academic_year,

            question_text=question.question_text,

        ).exists()

        if exists:

            duplicate_questions.append(
                question.question_text
            )

    # =====================================================
    # DUPLICATE QUESTIONS FOUND
    # =====================================================

    if duplicate_questions:

        print("\n========================================")
        print("DUPLICATE QUESTIONS")
        print("========================================")

        for duplicate in duplicate_questions:

            print(
                duplicate
            )

        messages.error(
            request,
            f"{len(duplicate_questions)} "
            f"question(s) already exist for "
            f"this course, semester, subject "
            f"and academic year."
        )

        messages.error(
            request,
            "Upload cancelled to prevent duplicate questions."
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # SAVE ALL QUESTIONS
    #
    # Atomic transaction:
    # Either ALL questions are saved,
    # or NONE are saved.
    # =====================================================

    try:

        with transaction.atomic():

            Question.objects.bulk_create(
                valid_questions
            )

    except Exception as e:

        print("\n========================================")
        print("DATABASE ERROR")
        print("========================================")

        print(e)

        messages.error(
            request,
            f"Unable to save questions: {e}"
        )

        return render(
            request,
            "admin_panel/upload_questions.html",
            {
                "form": form
            }
        )

    # =====================================================
    # SUCCESS
    # =====================================================

    created_count = len(
        valid_questions
    )

    print("\n========================================")
    print("UPLOAD SUCCESS")
    print("========================================")

    print(
        "Course:",
        course
    )

    print(
        "Semester:",
        semester
    )

    print(
        "Subject:",
        subject
    )

    print(
        "Academic Year:",
        academic_year
    )

    print(
        "Questions Created:",
        created_count
    )

    print("========================================\n")

    messages.success(
        request,
        f"{created_count} question(s) uploaded successfully."
    )

    return redirect(
        "question_bank"
    )
# PUBLISH EXAM
# =========================================

@login_required
@user_passes_test(is_admin)
def publish_exam(request, id):

    exam = get_object_or_404(
        Exam,
        id=id
    )


    exam.is_published = True

    exam.save()


    messages.success(
        request,
        f"{exam.exam_name} published successfully."
    )


    return redirect(
        "manage_exams"
    )

# =========================================
# DELETE QUESTION PAPER
# =========================================

@login_required
@user_passes_test(is_faculty)
def delete_paper(request, pk):

    paper = get_object_or_404(
        QuestionPaper,
        pk=pk
    )

    paper.delete()

    messages.success(
        request,
        "Question Paper deleted successfully."
    )

    return redirect("faculty_dashboard")

# =========================================
# ADMIN REPORTS DASHBOARD
# =========================================
@login_required
@user_passes_test(is_admin)
def admin_reports(request):

    results = Result.objects.select_related(
        "student",
        "student__studentprofile",
        "exam",
        "exam__course",
        "exam__subject"
    ).order_by("-completed_at")


    # ==========================================
    # FILTER VALUES
    # ==========================================

    search = request.GET.get(
        "search",
        ""
    ).strip()

    course = request.GET.get(
        "course",
        ""
    ).strip()

    semester = request.GET.get(
        "semester",
        ""
    ).strip()

    subject = request.GET.get(
        "subject",
        ""
    ).strip()

    exam_id = request.GET.get(
        "exam",
        ""
    ).strip()


    # ==========================================
    # SEARCH
    # ==========================================

    if search:

        results = results.filter(

            Q(student__username__icontains=search) |

            Q(student__first_name__icontains=search) |

            Q(student__last_name__icontains=search) |

            Q(student__email__icontains=search) |

            Q(exam__exam_name__icontains=search)

        )


    # ==========================================
    # COURSE
    # ==========================================

    if course:

        results = results.filter(
            exam__course_id=course
        )


    # ==========================================
    # SEMESTER
    # ==========================================

    if semester:

        try:

            semester_value = int(
                semester
            )

            if 1 <= semester_value <= 8:

                results = results.filter(
                    exam__semester=semester_value
                )

        except ValueError:

            semester = ""


    # ==========================================
    # SUBJECT
    # ==========================================

    if subject:

        results = results.filter(
            exam__subject_id=subject
        )


    # ==========================================
    # EXAM
    # ==========================================

    if exam_id:

        results = results.filter(
            exam_id=exam_id
        )


    # ==========================================
    # STATISTICS
    # ==========================================

    total_results = results.count()


    total_students = (
        results
        .values("student_id")
        .distinct()
        .count()
    )


    average_percentage = (
        results
        .aggregate(
            average=Avg("percentage")
        )["average"]
        or 0
    )


    pass_count = results.filter(
        percentage__gte=40
    ).count()


    fail_count = results.filter(
        percentage__lt=40
    ).count()


    highest_score = (
        results
        .order_by("-percentage")
        .first()
    )


    # ==========================================
    # FILTER OPTIONS
    # ==========================================

    courses = (
        Course.objects
        .all()
        .order_by("name")
    )


    subjects = (
        Subject.objects
        .all()
        .order_by("name")
    )


    exams = (
        Exam.objects
        .select_related(
            "course",
            "subject"
        )
        .order_by("-id")
    )


    semesters = range(1, 9)


    # ==========================================
    # CONTEXT
    # ==========================================

    context = {

        "results": results,

        "courses": courses,

        "subjects": subjects,

        "exams": exams,

        "semesters": semesters,


        "total_results":
            total_results,

        "total_students":
            total_students,

        "average_percentage":
            round(
                average_percentage,
                2
            ),

        "pass_count":
            pass_count,

        "fail_count":
            fail_count,

        "highest_score":
            highest_score,


        "search":
            search,

        "selected_course":
            course,

        "selected_semester":
            semester,

        "selected_subject":
            subject,

        "selected_exam":
            exam_id,

    }


    return render(
        request,
        "admin_panel/reports.html",
        context
    )
@login_required
@user_passes_test(is_admin)
def delete_exam_attempt(request, result_id):

    # =====================================================
    # ONLY POST REQUEST
    # =====================================================

    if request.method != "POST":
        return redirect("admin_reports")


    # =====================================================
    # GET RESULT
    # =====================================================

    result = get_object_or_404(
        Result.objects.select_related(
            "student",
            "exam"
        ),
        id=result_id
    )


    # =====================================================
    # SAVE INFORMATION BEFORE DELETE
    # =====================================================

    student = result.student
    exam = result.exam


    student_name = (
        student.get_full_name()
        or student.username
    )

    exam_name = exam.exam_name


    # =====================================================
    # DELETE STUDENT ANSWERS
    # =====================================================

    StudentAnswer.objects.filter(
        student=student,
        exam=exam
    ).delete()


    # =====================================================
    # DELETE RESULT
    # =====================================================

    result.delete()


    # =====================================================
    # SUCCESS MESSAGE
    # =====================================================

    messages.success(
        request,
        f"Previous attempt of {student_name} "
        f"for {exam_name} was deleted successfully. "
        f"The student can now attempt the exam again."
    )


    # =====================================================
    # PRESERVE REPORT FILTERS
    # =====================================================

    query_string = request.POST.get(
        "return_query",
        ""
    )


    if query_string:

        return redirect(
            f"/admin/reports/?{query_string}"
        )


    return redirect(
        "admin_reports"
    )    
# =========================================
# MANAGE EXAMS
# =========================================

@login_required
@user_passes_test(is_admin)
def manage_exams(request):

    exams = Exam.objects.select_related(
        "course",
        "subject"
    ).order_by("-created_at")

    return render(
        request,
        "admin_panel/manage_exams.html",
        {
            "exams": exams,
        }
    )




import random
import string


from .models import Course, StudentProfile


@login_required
@user_passes_test(is_admin_or_faculty)
def password_generator(request):

    courses = Course.objects.all().order_by("name")

    students = None
    generated_accounts = []

    selected_course = ""
    selected_semester = ""

    # ==========================================
    # SHOW STUDENTS - GET
    # ==========================================

    if request.method == "GET":

        selected_course = request.GET.get(
            "course",
            ""
        )

        selected_semester = request.GET.get(
            "semester",
            ""
        )

        if selected_course and selected_semester:

            students = (
                StudentProfile.objects
                .filter(
                    course_id=selected_course,
                    semester=selected_semester
                )
                .select_related(
                    "user",
                    "course"
                )
                .order_by(
                    "user__username"
                )
            )

    # ==========================================
    # GENERATE PASSWORDS - POST
    # ==========================================

    elif request.method == "POST":

        selected_course = request.POST.get(
            "course",
            ""
        )

        selected_semester = request.POST.get(
            "semester",
            ""
        )

        students = (
            StudentProfile.objects
            .filter(
                course_id=selected_course,
                semester=selected_semester
            )
            .select_related(
                "user",
                "course"
            )
            .order_by(
                "user__username"
            )
        )

        # Store passwords for Excel export
        generated_passwords = {}

        # ======================================
        # GENERATE PASSWORD FOR EACH STUDENT
        # ======================================

        for student in students:

            password = "".join(
                random.choices(
                    string.ascii_uppercase
                    + string.ascii_lowercase
                    + string.digits,
                    k=8
                )
            )

            # Save password securely
            student.user.set_password(
                password
            )

            student.user.save(
                update_fields=["password"]
            )

            # ----------------------------------
            # Store for display
            # ----------------------------------

            generated_accounts.append({

                "roll_no":
                    student.user.username,

                "name":
                    (
                        student.user.get_full_name().strip()
                        or student.user.username
                    ),

                "password":
                    password,
            })

            # ----------------------------------
            # IMPORTANT:
            # Store password for Excel export
            # ----------------------------------

            generated_passwords[
                str(student.id)
            ] = password

        # ======================================
        # SAVE GENERATED PASSWORDS IN SESSION
        # ======================================

        request.session[
            "generated_student_passwords"
        ] = generated_passwords

        request.session[
            "generated_password_course"
        ] = selected_course

        request.session[
            "generated_password_semester"
        ] = selected_semester

        request.session.modified = True

        # ======================================
        # SUCCESS MESSAGE
        # ======================================

        messages.success(
            request,
            f"{len(generated_accounts)} "
            f"passwords generated successfully."
        )

    # ==========================================
    # RENDER PAGE
    # ==========================================

    return render(
        request,
        "admin_panel/password_generator.html",
        {
            "courses":
                courses,

            "students":
                students,

            "generated_accounts":
                generated_accounts,

            "selected_course":
                selected_course,

            "selected_semester":
                selected_semester,

            # Used by template to show export button
            "passwords_generated":
                bool(
                    request.session.get(
                        "generated_student_passwords"
                    )
                ),
        }
    )
from django.http import HttpResponse
from openpyxl import Workbook


@login_required
@user_passes_test(is_admin_or_faculty)
def export_student_passwords(request):

    # ==========================================
    # GET SESSION DATA
    # ==========================================

    generated_passwords = request.session.get(
        "generated_student_passwords",
        {}
    )

    course_id = request.session.get(
        "generated_password_course"
    )

    semester = request.session.get(
        "generated_password_semester"
    )

    # ==========================================
    # VALIDATION
    # ==========================================

    if not generated_passwords:
        messages.error(
            request,
            "No generated passwords available for export."
        )

        return redirect(
            "password_generator"
        )

    if not course_id or not semester:
        messages.error(
            request,
            "Course and semester information is missing."
        )

        return redirect(
            "password_generator"
        )

    # ==========================================
    # GET COURSE
    # ==========================================

    course = get_object_or_404(
        Course,
        id=course_id
    )

    # ==========================================
    # GET STUDENTS
    # ==========================================

    students = (
        StudentProfile.objects
        .filter(
            id__in=generated_passwords.keys(),
            course_id=course_id,
            semester=semester
        )
        .select_related(
            "user",
            "course"
        )
        .order_by(
            "user__username"
        )
    )

    # ==========================================
    # CREATE EXCEL WORKBOOK
    # ==========================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Student Passwords"

    # ==========================================
    # HEADER
    # ==========================================

    worksheet.append([
        "Roll No",
        "Student Name",
        "Username",
        "Password",
        "Course",
        "Semester"
    ])

    # ==========================================
    # STUDENT DATA
    # ==========================================

    for student in students:

        user = student.user

        password = generated_passwords.get(
            str(student.id),
            ""
        )

        student_name = (
            user.get_full_name().strip()
            or user.username
        )

        worksheet.append([
            user.username,
            student_name,
            user.username,
            password,
            course.name,
            semester
        ])

    # ==========================================
    # COLUMN WIDTHS
    # ==========================================

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 30
    worksheet.column_dimensions["F"].width = 12

    # ==========================================
    # HTTP RESPONSE
    # ==========================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="student_passwords.xlsx"'
    )

    workbook.save(response)

    return response
@login_required
@user_passes_test(is_admin)
def admin_create_exam(request):

    if request.method == "POST":

        form = ExamForm(request.POST)

        if form.is_valid():

            exam = form.save(commit=False)

            exam.created_by = request.user

            # Admin-created exams start as draft
            exam.is_published = False

            # -----------------------------------------
            # GET AVAILABLE QUESTIONS
            # -----------------------------------------

            questions = list(
                Question.objects.filter(
                    course=exam.course,
                    subject=exam.subject,
                    semester=exam.semester,
                )
            )

            # -----------------------------------------
            # CHECK QUESTION COUNT
            # -----------------------------------------

            if len(questions) < exam.number_of_questions:

                messages.error(
                    request,
                    f"Only {len(questions)} questions are available "
                    f"for {exam.subject.name}, Semester {exam.semester}. "
                    f"You requested {exam.number_of_questions}."
                )

                return redirect("admin_create_exam")

            # -----------------------------------------
            # SAVE EXAM
            # -----------------------------------------

            exam.save()

            # -----------------------------------------
            # RANDOM QUESTIONS
            # -----------------------------------------

            selected_questions = random.sample(
                questions,
                exam.number_of_questions
            )

            # -----------------------------------------
            # CREATE EXAM QUESTIONS
            # -----------------------------------------

            ExamQuestion.objects.bulk_create(
                [
                    ExamQuestion(
                        exam=exam,
                        question=question
                    )
                    for question in selected_questions
                ]
            )

            messages.success(
                request,
                f"Exam '{exam.exam_name}' created successfully."
            )

            return redirect("manage_exams")

    else:

        form = ExamForm()

    # -----------------------------------------
    # ADMIN EXAM STATISTICS
    # -----------------------------------------

    exams = Exam.objects.all()

    context = {

        "form": form,

        "total_exams": exams.count(),

        "published_exams": exams.filter(
            is_published=True
        ).count(),

        "draft_exams": exams.filter(
            is_published=False
        ).count(),

        "total_questions": Question.objects.count(),

    }

    return render(
        request,
        "admin_panel/create_exam.html",
        context
    )

# =========================================================
# FACULTY - UPLOAD QUESTIONS
# =========================================================

# =========================================================
# FACULTY - UPLOAD QUESTIONS
# =========================================================

# =========================================================
# FACULTY - UPLOAD QUESTIONS
# =========================================================




# =========================================================
# FACULTY - UPLOAD QUESTIONS
# ============================================================
# IMPORTS
# ============================================================

import pandas as pd

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render, redirect

from .forms import QuestionUploadForm
from .models import Question, Subject


# ============================================================
# FACULTY CHECK
# ============================================================
@login_required
def faculty_upload_questions(request):

    # ---------------------------------------------------------
    # FACULTY ACCESS CHECK
    # ---------------------------------------------------------
    if not request.user.is_faculty and not request.user.is_superuser:
        messages.error(
            request,
            "You are not authorized to upload questions."
        )
        return redirect("faculty_login")

    # ---------------------------------------------------------
    # GET
    # ---------------------------------------------------------
    if request.method == "GET":

        form = FacultyQuestionUploadForm()

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    # ---------------------------------------------------------
    # POST
    # ---------------------------------------------------------
    form = FacultyQuestionUploadForm(request.POST, request.FILES)

    if not form.is_valid():

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    course = form.cleaned_data["course"]
    semester = form.cleaned_data["semester"]
    subject = form.cleaned_data["subject"]
    academic_year = form.cleaned_data["academic_year"].strip()
    excel_file = form.cleaned_data["excel_file"]

    print("=" * 70)
    print("FACULTY QUESTION UPLOAD")
    print("=" * 70)

    print("Course:", course)
    print("Semester:", semester)
    print("Subject:", subject)
    print("Academic Year:", academic_year)
    print("Excel:", excel_file.name)

    print("=" * 70)

    # ---------------------------------------------------------
    # READ EXCEL
    # ---------------------------------------------------------
    try:

        df = pd.read_excel(excel_file)

    except Exception as e:

        form.add_error(
            "excel_file",
            f"Unable to read Excel file: {e}"
        )

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    # ---------------------------------------------------------
    # CLEAN COLUMN NAMES
    # ---------------------------------------------------------
    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace("\n", " ", regex=False)
        .str.replace("\t", " ", regex=False)
    )

    # Fix accidental spaces between column names
    df.columns = [
        " ".join(column.split())
        for column in df.columns
    ]

    print("EXCEL COLUMNS:")
    print(list(df.columns))

    # ---------------------------------------------------------
    # EXPECTED COLUMNS
    # ---------------------------------------------------------
    required_columns = [
        "question_text",
        "option1",
        "option2",
        "option3",
        "option4",
        "correct_answer",
        "marks",
    ]

    # ---------------------------------------------------------
    # VALIDATE COLUMNS
    # ---------------------------------------------------------
    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:

        form.add_error(
            "excel_file",
            "Missing Excel columns: "
            + ", ".join(missing_columns)
        )

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    # ---------------------------------------------------------
    # REMOVE COMPLETELY EMPTY ROWS
    # ---------------------------------------------------------
    df = df.dropna(
        how="all"
    )

    print("NUMBER OF ROWS:")
    print(len(df))

    # ---------------------------------------------------------
    # IMPORTANT:
    # VALIDATE EVERYTHING BEFORE SAVING ANYTHING
    # ---------------------------------------------------------
    errors = []

    cleaned_rows = []

    for index, row in df.iterrows():

        excel_row = index + 2

        question_text = str(
            row["question_text"]
        ).strip()

        option1 = str(
            row["option1"]
        ).strip()

        option2 = str(
            row["option2"]
        ).strip()

        option3 = str(
            row["option3"]
        ).strip()

        option4 = str(
            row["option4"]
        ).strip()

        correct_answer = str(
            row["correct_answer"]
        ).strip()

        marks = row["marks"]

        print(
            f"Processing Excel row {excel_row}"
        )

        # -----------------------------------------------------
        # CHECK QUESTION
        # -----------------------------------------------------
        if (
            not question_text
            or question_text.lower() == "nan"
        ):

            errors.append(
                f"Excel row {excel_row}: "
                "Question is empty."
            )

            continue

        # -----------------------------------------------------
        # CHECK OPTIONS
        # -----------------------------------------------------
        options = {
            "Option 1": option1,
            "Option 2": option2,
            "Option 3": option3,
            "Option 4": option4,
        }

        row_has_error = False

        for option_name, option_value in options.items():

            if (
                not option_value
                or option_value.lower() == "nan"
            ):

                errors.append(
                    f"Excel row {excel_row}: "
                    f"{option_name} is empty."
                )

                row_has_error = True

        if row_has_error:
            continue

        # -----------------------------------------------------
        # CONVERT CORRECT ANSWER
        # -----------------------------------------------------
        answer_text = correct_answer.strip()

        answer_number = None

        # If Excel contains 1,2,3,4
        if answer_text in ["1", "2", "3", "4"]:

            answer_number = int(answer_text)

        else:

            answer_map = {
                option1.strip().lower(): 1,
                option2.strip().lower(): 2,
                option3.strip().lower(): 3,
                option4.strip().lower(): 4,
            }

            answer_number = answer_map.get(
                answer_text.lower()
            )

        # -----------------------------------------------------
        # INVALID CORRECT ANSWER
        # -----------------------------------------------------
        if answer_number is None:

            errors.append(
                f"Excel row {excel_row}: "
                f"Correct answer '{correct_answer}' "
                "does not match any option."
            )

            continue

        # -----------------------------------------------------
        # MARKS
        # -----------------------------------------------------
        try:

            marks_value = int(
                float(marks)
            )

        except (
            ValueError,
            TypeError
        ):

            errors.append(
                f"Excel row {excel_row}: "
                "Marks must be a number."
            )

            continue

        if marks_value <= 0:

            errors.append(
                f"Excel row {excel_row}: "
                "Marks must be greater than 0."
            )

            continue

        print(
            f"Correct answer converted to: "
            f"{answer_number}"
        )

        # -----------------------------------------------------
        # STORE CLEANED DATA
        # -----------------------------------------------------
        cleaned_rows.append(
            {
                "question_text": question_text,
                "option1": option1,
                "option2": option2,
                "option3": option3,
                "option4": option4,
                "correct_answer": answer_number,
                "marks": marks_value,
            }
        )

    # ---------------------------------------------------------
    # IF ANY ERROR, SAVE NOTHING
    # ---------------------------------------------------------
    if errors:

        print("=" * 70)
        print("UPLOAD VALIDATION ERROR")
        print("=" * 70)

        for error in errors:
            print(error)

        form.add_error(
            "excel_file",
            "Excel upload failed. "
            + " | ".join(errors)
        )

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    # ---------------------------------------------------------
    # NO VALID QUESTIONS
    # ---------------------------------------------------------
    if not cleaned_rows:

        form.add_error(
            "excel_file",
            "No valid questions found in the Excel file."
        )

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )

    # ---------------------------------------------------------
    # SAVE QUESTIONS
    # ---------------------------------------------------------
    created_count = 0

    try:

        for data in cleaned_rows:

            question = Question.objects.create(

                course=course,

                semester=int(
                    semester
                ),

                subject=subject,

                academic_year=academic_year,

                question_text=data[
                    "question_text"
                ],

                option1=data[
                    "option1"
                ],

                option2=data[
                    "option2"
                ],

                option3=data[
                    "option3"
                ],

                option4=data[
                    "option4"
                ],

                correct_answer=str(
                    data[
                        "correct_answer"
                    ]
                ),

                marks=data[
                    "marks"
                ],
            )

            created_count += 1

            print(
                f"CREATED QUESTION ID: "
                f"{question.id}"
            )

        print("=" * 70)
        print(
            f"SUCCESS: {created_count} "
            "questions uploaded."
        )
        print("=" * 70)

        messages.success(
            request,
            f"{created_count} questions uploaded successfully."
        )

        return redirect(
            "view_questions"
        )

    except Exception as e:

        print("=" * 70)
        print("DATABASE ERROR")
        print("=" * 70)
        print(e)

        form.add_error(
            "excel_file",
            f"Database error: {e}"
        )

        return render(
            request,
            "faculty/upload_questions.html",
            {
                "form": form
            }
        )
def is_faculty(user):
    return (
        user.is_authenticated
        and (
            getattr(user, "is_faculty", False)
            or user.is_superuser
        )
    )


# ============================================================
# GET SUBJECTS
# COURSE + SEMESTER
# ============================================================

@login_required
def get_subjects(request):

    course_id = request.GET.get("course")
    semester = request.GET.get("semester")

    if not course_id or not semester:
        return JsonResponse({
            "subjects": []
        })

    try:
        semester = int(semester)
    except (ValueError, TypeError):
        return JsonResponse({
            "subjects": []
        })

    subjects = Subject.objects.filter(
        course_id=course_id,
        semester=semester
    ).order_by("code")

    data = []

    for subject in subjects:

        data.append({
            "id": subject.id,
            "code": subject.code,
            "name": subject.name,
        })

    return JsonResponse({
        "subjects": data
    })


# ============================================================
# FACULTY - UPLOAD QUESTIONS
# ============================================================

import pandas as pd

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render, redirect

from .forms import QuestionUploadForm
from .models import Question, Subject
@login_required
@user_passes_test(is_admin)
def edit_exam(request, id):

    exam = get_object_or_404(
        Exam,
        id=id
    )

    if request.method == "POST":

        form = ExamForm(
            request.POST,
            instance=exam
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Exam updated successfully."
            )

            return redirect("manage_exams")

    else:

        form = ExamForm(
            instance=exam
        )

    return render(
        request,
        "admin_panel/edit_exam.html",
        {
            "form": form,
            "exam": exam
        }
    )




from .models import StudentProfile, Course


from django.db.models import Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test

from .models import StudentProfile, Course


@login_required
@user_passes_test(is_admin)
def manage_students(request):

    students = StudentProfile.objects.select_related(
        "user",
        "course"
    )

    # ==============================
    # SEARCH
    # ==============================

    search = request.GET.get("search", "").strip()

    if search:
        students = students.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )


    # ==============================
    # COURSE FILTER
    # ==============================

    course = request.GET.get("course", "").strip()

    if course:
        students = students.filter(
            course_id=course
        )


    # ==============================
    # SEMESTER FILTER
    # ==============================

    semester = request.GET.get("semester", "").strip()

    if semester:

        try:
            semester = int(semester)

            if 1 <= semester <= 8:
                students = students.filter(
                    semester=semester
                )

        except ValueError:
            pass


    # ==============================
    # ORDER
    # ==============================

    students = students.order_by("id")


    # ==============================
    # COURSES
    # ==============================

    courses = Course.objects.all().order_by("name")


    # ==============================
    # RENDER
    # ==============================

    return render(
        request,
        "admin_panel/manage_students.html",
        {
            "students": students,
            "courses": courses,
            "total_students": students.count(),
        }
    )

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

@login_required
def student_dashboard(request):

    # Prevent admin/faculty access
    if not request.user.is_student:
        messages.error(
            request,
            "Only students can access Student Dashboard."
        )

        if request.user.is_superuser:
            return redirect("admin_dashboard")

        if request.user.is_faculty:
            return redirect("faculty_dashboard")

        return redirect("index")

    student = get_object_or_404(
        StudentProfile,
        user=request.user
    )

    exams = Exam.objects.filter(
        course=student.course,
        is_published=True
    )

    results = Result.objects.filter(
        student=request.user
    )

    attempted_exam_ids = results.values_list(
        "exam_id",
        flat=True
    )

    context = {
        "student": student,
        "exams": exams,
        "results": results,
        "attempted_exam_ids": list(attempted_exam_ids),
    }

    return render(
        request,
        "student/dashboard.html",
        context
    )

from django.contrib.auth import logout
from django.shortcuts import redirect


def user_logout(request):

    logout(request)

    return redirect("index")

def load_subjects(request):

    course_id = request.GET.get("course")

    # your existing load_subjects code continues here

    subjects = Subject.objects.filter(
        course_id=course_id
    ).values(
        "id",
        "name",
        "semester"
    )

    return JsonResponse(
        list(subjects),
        safe=False
    )
@login_required
@user_passes_test(is_admin)
def view_exam_questions(request, exam_id):

    exam = get_object_or_404(
        Exam,
        id=exam_id
    )

    questions = ExamQuestion.objects.filter(
        exam=exam
    ).select_related(
        "question"
    )

    return render(
        request,
        "admin_panel/view_exam_questions.html",
        {
            "exam": exam,
            "questions": questions,
        }
    )
@login_required
@user_passes_test(is_admin)
def export_reports_pdf(request):

    from io import BytesIO

    from django.http import HttpResponse

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    # ==========================================
    # FILTER VALUES
    # ==========================================

    search = request.GET.get("search", "").strip()
    course = request.GET.get("course", "").strip()
    semester = request.GET.get("semester", "").strip()
    subject = request.GET.get("subject", "").strip()
    exam_id = request.GET.get("exam", "").strip()

    # ==========================================
    # RESULTS
    # ==========================================

    results = Result.objects.select_related(
        "student",
        "student__studentprofile",
        "exam",
        "exam__course",
        "exam__subject"
    ).order_by("-completed_at")

    # ==========================================
    # APPLY FILTERS
    # ==========================================

    if search:
        results = results.filter(
            Q(student__username__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__email__icontains=search) |
            Q(exam__exam_name__icontains=search)
        )

    if course:
        results = results.filter(
            exam__course_id=course
        )

    if semester:
        try:
            semester_value = int(semester)

            if 1 <= semester_value <= 8:
                results = results.filter(
                    exam__semester=semester_value
                )

        except ValueError:
            pass

    if subject:
        results = results.filter(
            exam__subject_id=subject
        )

    if exam_id:
        results = results.filter(
            exam_id=exam_id
        )

    # ==========================================
    # PDF
    # ==========================================

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b>CIA Examination Management System</b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Student Examination Reports",
            styles["Heading2"]
        )
    )

    elements.append(
        Spacer(1, 8)
    )

    # ==========================================
    # FILTER SUMMARY
    # ==========================================

    filter_parts = []

    if course:
        try:
            course_obj = Course.objects.get(
                id=course
            )
            filter_parts.append(
                f"Course: {course_obj.name}"
            )
        except Course.DoesNotExist:
            pass

    if semester:
        filter_parts.append(
            f"Semester: {semester}"
        )

    if subject:
        try:
            subject_obj = Subject.objects.get(
                id=subject
            )
            filter_parts.append(
                f"Subject: {subject_obj.name}"
            )
        except Subject.DoesNotExist:
            pass

    if exam_id:
        try:
            exam_obj = Exam.objects.get(
                id=exam_id
            )
            filter_parts.append(
                f"Exam: {exam_obj.exam_name}"
            )
        except Exam.DoesNotExist:
            pass

    if search:
        filter_parts.append(
            f"Search: {search}"
        )

    if filter_parts:
        filter_text = (
            "<b>Filters:</b> "
            + " | ".join(filter_parts)
        )
    else:
        filter_text = (
            "<b>Filters:</b> All Results"
        )

    elements.append(
        Paragraph(
            filter_text,
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    # ==========================================
    # TABLE
    # ==========================================

    data = [
        [
            "Student",
            "Username",
            "Exam",
            "Course",
            "Sem",
            "Subject",
            "Score",
            "Total",
            "%",
            "Status",
            "Date",
        ]
    ]

    for result in results:

        student_name = result.student.get_full_name()

        if not student_name:
            try:
                student_name = (
                    result.student.studentprofile.name
                )
            except StudentProfile.DoesNotExist:
                student_name = result.student.username

        status = (
            "PASS"
            if result.percentage >= 40
            else "FAIL"
        )

        data.append([
            student_name,
            result.student.username,
            result.exam.exam_name,
            result.exam.course.name
            if result.exam.course else "",
            result.exam.semester,
            result.exam.subject.name
            if result.exam.subject else "",
            result.score,
            result.total_marks,
            f"{result.percentage:.2f}",
            status,
            result.completed_at.strftime(
                "%d-%m-%Y"
            ),
        ])

    if len(data) == 1:

        data.append([
            "No results found",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            30 * mm,
            25 * mm,
            28 * mm,
            25 * mm,
            12 * mm,
            28 * mm,
            13 * mm,
            13 * mm,
            13 * mm,
            16 * mm,
            23 * mm,
        ]
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#212529"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE",
            ),
            (
                "ALIGN",
                (4, 1),
                (9, -1),
                "CENTER",
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f8f9fa"),
                ],
            ),
            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                4,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                4,
            ),
        ])
    )

    elements.append(table)

    document.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    response = HttpResponse(
        pdf,
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="exam_reports.pdf"'
    )

    return response
@login_required
@user_passes_test(is_admin)
def export_reports_excel(request):

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse

    # ==========================================
    # FILTER VALUES
    # ==========================================

    search = request.GET.get("search", "").strip()
    course = request.GET.get("course", "").strip()
    semester = request.GET.get("semester", "").strip()
    subject = request.GET.get("subject", "").strip()
    exam_id = request.GET.get("exam", "").strip()

    # ==========================================
    # RESULTS
    # ==========================================

    results = Result.objects.select_related(
        "student",
        "student__studentprofile",
        "exam",
        "exam__course",
        "exam__subject"
    ).order_by("-completed_at")

    # ==========================================
    # APPLY FILTERS
    # ==========================================

    if search:
        results = results.filter(
            Q(student__username__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__email__icontains=search) |
            Q(exam__exam_name__icontains=search)
        )

    if course:
        results = results.filter(
            exam__course_id=course
        )

    if semester:
        try:
            semester_value = int(semester)

            if 1 <= semester_value <= 8:
                results = results.filter(
                    exam__semester=semester_value
                )

        except ValueError:
            pass

    if subject:
        results = results.filter(
            exam__subject_id=subject
        )

    if exam_id:
        results = results.filter(
            exam_id=exam_id
        )

    # ==========================================
    # CREATE WORKBOOK
    # ==========================================

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Exam Reports"

    headers = [
        "Student Name",
        "Username",
        "Exam",
        "Course",
        "Semester",
        "Subject",
        "Score",
        "Total Marks",
        "Percentage",
        "Status",
        "Completed Date",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # DATA
    # ==========================================

    for result in results:

        student_name = result.student.get_full_name()

        if not student_name:
            try:
                student_name = (
                    result.student.studentprofile.name
                )
            except StudentProfile.DoesNotExist:
                student_name = result.student.username

        status = (
            "PASS"
            if result.percentage >= 40
            else "FAIL"
        )

        worksheet.append([
            student_name,
            result.student.username,
            result.exam.exam_name,
            result.exam.course.name
            if result.exam.course else "",
            result.exam.semester,
            result.exam.subject.name
            if result.exam.subject else "",
            result.score,
            result.total_marks,
            result.percentage,
            status,
            result.completed_at.strftime(
                "%d-%m-%Y %H:%M"
            ),
        ])

    # ==========================================
    # COLUMN WIDTHS
    # ==========================================

    widths = {
        "A": 25,
        "B": 20,
        "C": 25,
        "D": 25,
        "E": 12,
        "F": 25,
        "G": 12,
        "H": 15,
        "I": 15,
        "J": 12,
        "K": 22,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    # ==========================================
    # RESPONSE
    # ==========================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="exam_reports.xlsx"'
    )

    workbook.save(response)

    return response
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test


@login_required
@user_passes_test(is_faculty)
def export_results(request):

    results = Result.objects.select_related(
        "student",
        "exam",
        "exam__course",
        "exam__subject"
    )


    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Student Results"


    sheet.append([
        "Student",
        "Username",
        "Exam",
        "Course",
        "Subject",
        "Score",
        "Total Marks",
        "Percentage",
        "Status",
        "Date"
    ])


    for result in results:

        status = "PASS" if result.percentage >= 40 else "FAIL"


        sheet.append([

            result.student.studentprofile.name,

            result.student.username,

            result.exam.exam_name,

            result.exam.course.name,

            result.exam.subject.name,

            result.score,

            result.total_marks,

            result.percentage,

            status,

            result.completed_at.strftime("%d-%m-%Y %H:%M")

        ])


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="student_results.xlsx"'
    )


    workbook.save(response)


    return response
@login_required
@user_passes_test(is_faculty)
def faculty_edit_question(request, pk):

    question = get_object_or_404(
        Question,
        pk=pk
    )

    if request.method == "POST":

        form = QuestionForm(
            request.POST,
            instance=question
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Question updated successfully."
            )

            return redirect("view_questions")

    else:

        form = QuestionForm(
            instance=question
        )

    return render(
        request,
        "faculty/edit_question.html",
        {
            "form": form,
            "question": question,
        }
    )
from django.http import FileResponse
from django.conf import settings
import os


@login_required
@user_passes_test(is_faculty)
def download_question_template(request):
    file_path = os.path.join(
        settings.MEDIA_ROOT,
        "question_template.xlsx"
    )

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename="question_template.xlsx"
    )
@login_required
@user_passes_test(is_admin)
def published_exams(request):

    exams = Exam.objects.filter(
        is_published=True
    ).select_related(
        "course",
        "subject"
    ).order_by("-id")


    return render(
        request,
        "admin_panel/published_exams.html",
        {
            "exams": exams
        }
    )
@login_required
@user_passes_test(is_admin)
def unpublish_exam(request, id):

    exam = get_object_or_404(
        Exam,
        id=id
    )

    exam.is_published = False

    exam.save()


    messages.success(
        request,
        "Exam unpublished successfully."
    )


    return redirect(
        "published_exams"
    )
@login_required
@user_passes_test(lambda u: u.is_student)
def student_upcoming_exams(request):

    student = get_object_or_404(
        StudentProfile,
        user=request.user
    )


    exams = Exam.objects.filter(
        course=student.course,
        is_published=True
    )


    return render(
        request,
        "student/upcoming_exams.html",
        {
            "student": student,
            "exams": exams
        }
    )



@login_required
@user_passes_test(lambda u: u.is_student)
def student_results(request):

    results = Result.objects.filter(
        student=request.user
    ).select_related(
        "exam",
        "exam__subject"
    ).order_by(
        "-completed_at"
    )


    return render(
        request,
        "student/result.html",
        {
            "results": results
        }
    )
@login_required
@user_passes_test(is_faculty)
def delete_all_questions(request):

    if request.method == "POST":

        count = Question.objects.count()

        Question.objects.all().delete()

        messages.success(
            request,
            f"{count} questions deleted successfully."
        )

    return redirect("view_questions")
@login_required
@user_passes_test(is_admin)
def edit_student(request, id):

    student = get_object_or_404(
        StudentProfile,
        id=id
    )

    if request.method == "POST":

        form = StudentRegistrationForm(
            request.POST,
            instance=student.user
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student updated successfully."
            )

            return redirect(
                "manage_students"
            )

    else:

        form = StudentRegistrationForm(
            instance=student.user
        )


    return render(
        request,
        "admin_panel/edit_student.html",
        {
            "form": form,
            "student": student
        }
    )
@login_required
@user_passes_test(is_admin)
def delete_selected_students(request):

    if request.method == "POST":

        student_ids = request.POST.getlist("student_ids")

        if student_ids:

            students = StudentProfile.objects.filter(
                id__in=student_ids
            )

            for student in students:
                student.user.delete()

            messages.success(
                request,
                f"{len(student_ids)} students deleted successfully."
            )

        else:

            messages.warning(
                request,
                "Please select students to delete."
            )


    return redirect("manage_students")


User = get_user_model()
@login_required
@user_passes_test(is_admin)
def upload_faculty(request):

    courses = Course.objects.all().order_by("name")
    subjects = Subject.objects.all().order_by("name")

    template_name = "admin_panel/upload_faculty.html"

    if request.method == "GET":
        return render(
            request,
            template_name,
            {
                "courses": courses,
                "subjects": subjects,
            }
        )

    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        messages.error(
            request,
            "Please select an Excel file."
        )

        return render(
            request,
            template_name,
            {
                "courses": courses,
                "subjects": subjects,
            }
        )

    if not excel_file.name.lower().endswith(
        (".xlsx", ".xls")
    ):
        messages.error(
            request,
            "Only .xlsx and .xls files are allowed."
        )

        return render(
            request,
            template_name,
            {
                "courses": courses,
                "subjects": subjects,
            }
        )

    try:
        df = pd.read_excel(
            excel_file,
            sheet_name=0,
            dtype=str
        )
    except Exception as e:
        messages.error(
            request,
            f"Unable to read Excel file: {e}"
        )

        return render(
            request,
            template_name,
            {
                "courses": courses,
                "subjects": subjects,
            }
        )

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
    )

    df = df.dropna(how="all")

    required_columns = [
        "username",
        "name",
        "first name",
        "last name",
        "email",
        "department",
        "course",
        "subjects",
        "password",
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        messages.error(
            request,
            "Missing Excel columns: "
            + ", ".join(missing_columns)
        )

        return render(
            request,
            template_name,
            {
                "courses": courses,
                "subjects": subjects,
            }
        )

    created_count = 0
    skipped_count = 0
    errors = []

    print("======================================")
    print("FACULTY BULK UPLOAD")
    print("COLUMNS:", list(df.columns))
    print("ROWS:", len(df))
    print("======================================")

    for row_number, row in enumerate(
        df.to_dict("records"),
        start=2
    ):

        username = str(
            row["username"]
        ).strip()

        name = str(
            row["name"]
        ).strip()

        first_name = str(
            row["first name"]
        ).strip()

        last_name = str(
            row["last name"]
        ).strip()

        email = str(
            row["email"]
        ).strip()

        department = str(
            row["department"]
        ).strip()

        course_name = str(
            row["course"]
        ).strip()

        subjects_value = str(
            row["subjects"]
        ).strip()

        password = str(
            row["password"]
        ).strip()

        # Clean NaN values
        if username.lower() == "nan":
            username = ""

        if name.lower() == "nan":
            name = ""

        if first_name.lower() == "nan":
            first_name = ""

        if last_name.lower() == "nan":
            last_name = ""

        if email.lower() == "nan":
            email = ""

        if department.lower() == "nan":
            department = ""

        if course_name.lower() == "nan":
            course_name = ""

        if subjects_value.lower() == "nan":
            subjects_value = ""

        if password.lower() == "nan":
            password = ""

        # Required validation
        if not username:
            errors.append(
                f"Row {row_number}: Username is required."
            )
            skipped_count += 1
            continue

        if not name:
            errors.append(
                f"Row {row_number}: Name is required."
            )
            skipped_count += 1
            continue

        if not email:
            errors.append(
                f"Row {row_number}: Email is required."
            )
            skipped_count += 1
            continue

        if not department:
            errors.append(
                f"Row {row_number}: Department is required."
            )
            skipped_count += 1
            continue

        if not course_name:
            errors.append(
                f"Row {row_number}: Course is required."
            )
            skipped_count += 1
            continue

        # Username duplicate
        if User.objects.filter(
            username=username
        ).exists():

            errors.append(
                f"Row {row_number}: "
                f"Username '{username}' already exists."
            )

            skipped_count += 1
            continue

        # Course
        course = Course.objects.filter(
            name__iexact=course_name
        ).first()

        if course is None:

            errors.append(
                f"Row {row_number}: "
                f"Course '{course_name}' does not exist."
            )

            skipped_count += 1
            continue

        # Subjects
        subject_objects = []
        invalid_subjects = []

        subject_names = [
            value.strip()
            for value in subjects_value.split(",")
            if value.strip()
        ]

        for subject_name in subject_names:

            subject = Subject.objects.filter(
                name__iexact=subject_name
            ).first()

            if subject:
                subject_objects.append(subject)
            else:
                invalid_subjects.append(
                    subject_name
                )

        if invalid_subjects:

            errors.append(
                f"Row {row_number}: "
                f"Subject(s) not found: "
                f"{', '.join(invalid_subjects)}"
            )

            skipped_count += 1
            continue

        # Default password
        if not password:
            password = f"CIA@{username}"

        user = None

        try:

            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_faculty=True,
            )

            faculty = FacultyProfile.objects.create(
                user=user,
                department=department,
                course=course,
            )

            if subject_objects:
                faculty.subjects.set(
                    subject_objects
                )

            created_count += 1

            print(
                f"CREATED FACULTY: {username}"
            )

        except Exception as e:

            if user is not None:
                user.delete()

            errors.append(
                f"Row {row_number}: {str(e)}"
            )

            skipped_count += 1

    print("======================================")
    print("FACULTY UPLOAD COMPLETE")
    print("CREATED:", created_count)
    print("SKIPPED:", skipped_count)
    print("ERRORS:", errors)
    print("======================================")

    if created_count:
        messages.success(
            request,
            f"{created_count} faculty member(s) "
            f"uploaded successfully."
        )

    if skipped_count:
        messages.warning(
            request,
            f"{skipped_count} row(s) were skipped."
        )

    for error in errors:
        messages.error(
            request,
            error
        )

    return render(
        request,
        template_name,
        {
            "courses": courses,
            "subjects": subjects,
        }
    )


@login_required
@user_passes_test(is_admin)
def download_faculty_template(request):

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Faculty Template"

    # ==========================================
    # HEADERS
    # ==========================================

    headers = [
        "Username",
        "Name",
        "First Name",
        "Last Name",
        "Email",
        "Department",
        "Course",
        "Subjects",
        "Password",
    ]

    worksheet.append(headers)

    # ==========================================
    # HEADER STYLE
    # ==========================================

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================
    # SAMPLE ROW
    # ==========================================

    worksheet.append([
        "faculty101",
        "Ravi Kumar",
        "Ravi",
        "Kumar",
        "ravi@gmail.com",
        "Computer Science",
        "BCA",
        "java,C-Language",
        "CIA@faculty101",
    ])

    # ==========================================
    # COLUMN WIDTHS
    # ==========================================

    widths = {
        "A": 20,
        "B": 25,
        "C": 20,
        "D": 20,
        "E": 30,
        "F": 25,
        "G": 25,
        "H": 35,
        "I": 22,
    }

    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # ==========================================
    # RESPONSE
    # ==========================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="faculty_upload_template.xlsx"'
    )

    workbook.save(response)

    return response
@login_required
@user_passes_test(is_admin)
def delete_exam(request, exam_id):

    # =====================================================
    # ONLY POST REQUEST
    # =====================================================

    if request.method != "POST":
        return redirect("manage_exams")


    # =====================================================
    # GET EXAM
    # =====================================================

    exam = get_object_or_404(
        Exam,
        id=exam_id
    )


    # =====================================================
    # SAVE EXAM NAME
    # =====================================================

    exam_name = exam.exam_name


    # =====================================================
    # DELETE EVERYTHING RELATED TO THIS EXAM
    # =====================================================

    with transaction.atomic():

        # -------------------------------------------------
        # DELETE STUDENT ANSWERS
        # -------------------------------------------------

        StudentAnswer.objects.filter(
            exam=exam
        ).delete()


        # -------------------------------------------------
        # DELETE RESULTS
        # -------------------------------------------------

        Result.objects.filter(
            exam=exam
        ).delete()


        # -------------------------------------------------
        # DELETE EXAM QUESTIONS
        # -------------------------------------------------

        ExamQuestion.objects.filter(
            exam=exam
        ).delete()


        # -------------------------------------------------
        # DELETE EXAM
        # -------------------------------------------------

        exam.delete()


    # =====================================================
    # SUCCESS MESSAGE
    # =====================================================

    messages.success(
        request,
        f'Exam "{exam_name}" deleted successfully.'
    )


    # =====================================================
    # RETURN TO EXAM MANAGEMENT
    # =====================================================

    return redirect(
        "manage_exams"
    )

@login_required
@user_passes_test(is_admin)
def bulk_delete_students(request):

    if request.method != "POST":
        return redirect("manage_students")

    student_ids = request.POST.getlist("student_ids")

    if not student_ids:
        messages.warning(
            request,
            "Please select at least one student to delete."
        )
        return redirect("manage_students")

    students = StudentProfile.objects.filter(
        id__in=student_ids
    ).select_related("user")

    deleted_count = 0

    for student in students:

        user = student.user

        # Delete StudentProfile first
        student.delete()

        # Delete associated login account
        if user:
            user.delete()

        deleted_count += 1

    messages.success(
        request,
        f"{deleted_count} student(s) deleted successfully."
    )

    return redirect("manage_students")


@login_required
@user_passes_test(is_admin_or_faculty)
def export_student_passwords(request):

    # ==========================================
    # GET GENERATED PASSWORDS FROM SESSION
    # ==========================================

    generated_passwords = request.session.get(
        "generated_student_passwords",
        {}
    )

    # ==========================================
    # NOTHING TO EXPORT
    # ==========================================

    if not generated_passwords:

        messages.warning(
            request,
            "No generated passwords are available for export."
        )

        return redirect(
            request.META.get(
                "HTTP_REFERER",
                "/faculty/password-generator/"
            )
        )

    # ==========================================
    # CREATE EXCEL
    # ==========================================

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Student Passwords"

    # ==========================================
    # HEADERS
    # ==========================================

    headers = [
        "Roll No",
        "Student Name",
        "Username",
        "Email",
        "Course",
        "Semester",
        "Generated Password",
    ]

    worksheet.append(headers)

    # ==========================================
    # HEADER STYLE
    # ==========================================

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==========================================
    # GET STUDENT IDs
    # ==========================================

    student_ids = list(
        generated_passwords.keys()
    )

    # ==========================================
    # GET STUDENTS
    # ==========================================

    students = (
        StudentProfile.objects
        .filter(
            id__in=student_ids
        )
        .select_related(
            "user",
            "course"
        )
        .order_by(
            "user__username"
        )
    )

    # ==========================================
    # ADD STUDENTS
    # ==========================================

    for student in students:

        password = generated_passwords.get(
            str(student.id),
            ""
        )

        # Student name
        student_name = (
            student.user.get_full_name().strip()
        )

        if not student_name:

            student_name = (
                getattr(
                    student,
                    "name",
                    ""
                )
                or student.user.username
            )

        # Course
        if student.course:

            course_name = (
                student.course.name
            )

        else:

            course_name = "Not Assigned"

        # Semester
        if student.semester:

            semester_name = (
                f"Semester {student.semester}"
            )

        else:

            semester_name = "Not Assigned"

        # ======================================
        # EXCEL ROW
        # ======================================

        worksheet.append([
            student.user.username,
            student_name,
            student.user.username,
            student.user.email or "Not Available",
            course_name,
            semester_name,
            password,
        ])

    # ==========================================
    # FORMAT COLUMNS
    # ==========================================

    column_widths = {

        "A": 22,

        "B": 40,

        "C": 22,

        "D": 40,

        "E": 20,

        "F": 18,

        "G": 25,
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # ==========================================
    # FREEZE HEADER
    # ==========================================

    worksheet.freeze_panes = "A2"

    # ==========================================
    # FILTER
    # ==========================================

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # ==========================================
    # ALIGNMENT
    # ==========================================

    for row in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="center"
            )

    # ==========================================
    # CREATE DOWNLOAD
    # ==========================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="student_passwords.xlsx"'
    )

    return response
# =========================================
# GET SUBJECTS BY COURSE + SEMESTER
# =========================================



# =========================================================
# GET SUBJECTS BY COURSE AND SEMESTER
# =========================================================

@login_required
@user_passes_test(is_admin_or_faculty)
def get_subjects_by_course_semester(request):

    course_id = request.GET.get("course")
    semester = request.GET.get("semester")

    if not course_id or not semester:

        return JsonResponse({
            "subjects": []
        })

    subjects = Subject.objects.filter(
        course_id=course_id,
        semester=semester
    ).order_by(
        "code",
        "name"
    )

    data = []

    for subject in subjects:

        data.append({
            "id": subject.id,
            "code": subject.code,
            "name": subject.name,
        })

    return JsonResponse({
        "subjects": data
    })
@login_required
def load_subjects(request):

    course_id = request.GET.get("course")
    semester = request.GET.get("semester")

    if not course_id or not semester:
        return JsonResponse(
            {
                "subjects": []
            }
        )

    try:

        course_id = int(course_id)
        semester = int(semester)

    except (ValueError, TypeError):

        return JsonResponse(
            {
                "subjects": []
            },
            status=400
        )

    subjects = Subject.objects.filter(
        course_id=course_id,
        semester=semester
    ).order_by("code")

    data = []

    for subject in subjects:

        data.append(
            {
                "id": subject.id,
                "code": subject.code,
                "name": subject.name,
            }
        )

    return JsonResponse(
        {
            "subjects": data
        }
    )




# =========================================================
# GET SUBJECTS BY COURSE AND SEMESTER
# =========================================================

@login_required
@user_passes_test(is_admin_or_faculty)
def get_subjects(request):

    # =====================================================
    # GET COURSE AND SEMESTER
    # =====================================================

    course_id = request.GET.get("course")
    semester = request.GET.get("semester")


    # =====================================================
    # VALIDATE INPUT
    # =====================================================

    if not course_id or not semester:

        return JsonResponse({
            "subjects": []
        })


    # =====================================================
    # CONVERT TO INTEGER
    # =====================================================

    try:

        course_id = int(course_id)
        semester = int(semester)

    except (ValueError, TypeError):

        return JsonResponse({
            "subjects": []
        })


    # =====================================================
    # GET SUBJECTS
    # =====================================================

    subjects = Subject.objects.filter(
        course_id=course_id,
        semester=semester
    ).order_by(
        "code",
        "name"
    )


    # =====================================================
    # PREPARE JSON DATA
    # =====================================================

    data = []

    for subject in subjects:

        data.append({

            "id": subject.id,

            "code": subject.code,

            "name": subject.name,

        })


    # =====================================================
    # RETURN JSON
    # =====================================================

    return JsonResponse({

        "subjects": data

    })
@login_required
@user_passes_test(is_admin_or_faculty)
def upload_question_bank(request):

    if request.method == "POST":

        form = QuestionUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            course = form.cleaned_data["course"]
            semester = form.cleaned_data["semester"]
            subject = form.cleaned_data["subject"]
            academic_year = form.cleaned_data["academic_year"]
            excel_file = form.cleaned_data["excel_file"]

            # ==================================================
            # YOUR EXISTING EXCEL PROCESSING CODE
            # ==================================================
            #
            # Keep your current pandas/openpyxl Question
            # creation logic here.
            #
            # Do NOT replace that logic if it is already working.
            #
            # ==================================================

            messages.success(
                request,
                "Questions uploaded successfully."
            )

            return redirect(
                "question_bank"
            )

    else:

        form = QuestionUploadForm()

    return render(
        request,
        "admin_panel/upload_question_bank.html",
        {
            "form": form
        }
    )
@login_required
@user_passes_test(is_admin_or_faculty)
def delete_question(request, pk):

    # =====================================================
    # GET QUESTION
    # =====================================================

    question = get_object_or_404(
        Question,
        pk=pk
    )

    # =====================================================
    # DELETE ONLY WITH POST
    # =====================================================

    if request.method == "POST":

        question.delete()

        messages.success(
            request,
            "Question deleted successfully."
        )

        # =================================================
        # ADMIN
        # =================================================

        if request.user.is_superuser:

            return redirect(
                "question_bank"
            )

        # =================================================
        # FACULTY
        # =================================================

        return redirect(
            "view_questions"
        )

    # =====================================================
    # GET REQUEST
    # DO NOT DELETE
    # =====================================================

    if request.user.is_superuser:

        return redirect(
            "question_bank"
        )

    return redirect(
        "view_questions"
    )@login_required
@user_passes_test(is_faculty)
def faculty_upload_questions(request):

    if request.method == "POST":

        form = QuestionUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            course = form.cleaned_data["course"]
            semester = form.cleaned_data["semester"]
            subject = form.cleaned_data["subject"]
            academic_year = form.cleaned_data["academic_year"]
            excel_file = form.cleaned_data["excel_file"]

            try:

                import pandas as pd

                df = pd.read_excel(excel_file)

                df.columns = (
                    df.columns
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )

                required_columns = [

                    "question_text",
                    "option1",
                    "option2",
                    "option3",
                    "option4",
                    "correct_answer",
                    "marks",

                ]

                missing_columns = [

                    column
                    for column in required_columns
                    if column not in df.columns

                ]

                if missing_columns:

                    messages.error(
                        request,
                        "Missing Excel columns: "
                        + ", ".join(missing_columns)
                    )

                    return render(
                        request,
                        "faculty/faculty_upload_questions.html",
                        {
                            "form": form,
                        }
                    )

                created_count = 0

                for index, row in df.iterrows():

                    question_text = str(
                        row["question_text"]
                    ).strip()

                    if not question_text:

                        continue

                    correct_answer = str(
                        row["correct_answer"]
                    ).strip()

                    if correct_answer.endswith(".0"):

                        correct_answer = correct_answer[:-2]

                    if correct_answer not in [
                        "1",
                        "2",
                        "3",
                        "4",
                    ]:

                        continue

                    def clean_value(value):

                        if pd.isna(value):

                            return ""

                        return str(value).strip()


                    Question.objects.create(

                        question_text=question_text,

                        option1=clean_value(
                            row["option1"]
                        ),

                        option2=clean_value(
                            row["option2"]
                        ),

                        option3=clean_value(
                            row["option3"]
                        ),

                        option4=clean_value(
                            row["option4"]
                        ),

                        correct_answer=correct_answer,

                        marks=int(
                            row["marks"]
                        ),

                        course=course,

                        semester=semester,

                        subject=subject,

                        academic_year=academic_year,

                    )

                    created_count += 1


                messages.success(

                    request,

                    f"{created_count} questions uploaded successfully."

                )

                return redirect(
                    "faculty_upload_questions"
                )


            except Exception as e:

                messages.error(

                    request,

                    f"Error uploading Excel file: {str(e)}"

                )


    else:

        form = QuestionUploadForm()


    return render(

        request,

        "faculty/faculty_upload_questions.html",

        {
            "form": form,
        }

    )
@login_required
@user_passes_test(is_admin)
def add_student(request):

    if request.method == "POST":

        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        course_id = request.POST.get("course")
        semester = request.POST.get("semester")

        if not username or not password:
            messages.error(
                request,
                "Username and password are required."
            )
            return redirect("add_student")

        if User.objects.filter(username=username).exists():
            messages.error(
                request,
                "Username already exists."
            )
            return redirect("add_student")

        course = None

        if course_id:
            course = Course.objects.filter(
                id=course_id
            ).first()

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            last_name=last_name,
            email=email,
            is_student=True,
            semester=int(semester) if semester else 1,
            course=course
        )

        StudentProfile.objects.create(
            user=user
        )

        messages.success(
            request,
            f"Student {username} created successfully."
        )

        return redirect("manage_students")

    courses = Course.objects.all().order_by("name")

    return render(
        request,
        "admin_panel/add_student.html",
        {
            "courses": courses
        }
    )
# =========================================================
# EDIT QUESTION
# =========================================================

@login_required
def edit_question(request, pk):

    question = get_object_or_404(
        Question,
        pk=pk
    )

    # =====================================================
    # ACCESS CONTROL
    # =====================================================

    if request.user.is_superuser:

        # Admin can edit any question
        pass

    elif request.user.is_faculty:

        # Faculty can edit questions
        pass

    else:

        messages.error(
            request,
            "You are not allowed to edit questions."
        )

        return redirect("student_dashboard")


    # =====================================================
    # RETURN URL
    # =====================================================

    next_url = request.POST.get(
        "next"
    ) or request.GET.get(
        "next"
    )


    # =====================================================
    # POST
    # =====================================================

    if request.method == "POST":

        question_text = request.POST.get(
            "question_text",
            ""
        ).strip()

        option1 = request.POST.get(
            "option1",
            ""
        ).strip()

        option2 = request.POST.get(
            "option2",
            ""
        ).strip()

        option3 = request.POST.get(
            "option3",
            ""
        ).strip()

        option4 = request.POST.get(
            "option4",
            ""
        ).strip()

        correct_answer = request.POST.get(
            "correct_answer",
            ""
        ).strip()

        marks = request.POST.get(
            "marks",
            ""
        ).strip()


        # =================================================
        # VALIDATION
        # =================================================

        if not question_text:

            messages.error(
                request,
                "Question text cannot be empty."
            )

            return render(
                request,
                "admin_panel/edit_question.html",
                {
                    "question": question,
                    "next_url": next_url,
                }
            )


        if correct_answer not in [
            "1",
            "2",
            "3",
            "4"
        ]:

            messages.error(
                request,
                "Please select a valid correct answer."
            )

            return render(
                request,
                "admin_panel/edit_question.html",
                {
                    "question": question,
                    "next_url": next_url,
                }
            )


        try:

            marks_value = int(marks)

            if marks_value <= 0:

                raise ValueError

        except (ValueError, TypeError):

            messages.error(
                request,
                "Marks must be a positive number."
            )

            return render(
                request,
                "admin_panel/edit_question.html",
                {
                    "question": question,
                    "next_url": next_url,
                }
            )


        # =================================================
        # UPDATE QUESTION
        # =================================================

        question.question_text = question_text

        question.option1 = option1

        question.option2 = option2

        question.option3 = option3

        question.option4 = option4

        question.correct_answer = correct_answer

        question.marks = marks_value


        # =================================================
        # SAVE
        # =================================================

        question.save()


        messages.success(
            request,
            "Question updated successfully."
        )


        # =================================================
        # RETURN TO PREVIOUS QUESTION BANK
        # =================================================

        if next_url:

            if request.user.is_superuser:

                return redirect(
                    f"{next_url}"
                )

            elif request.user.is_faculty:

                return redirect(
                    f"{next_url}"
                )


        # =================================================
        # DEFAULT REDIRECT
        # =================================================

        if request.user.is_superuser:

            return redirect(
                "question_bank"
            )

        elif request.user.is_faculty:

            return redirect(
                "faculty_question_bank"
            )


    # =====================================================
    # GET
    # =====================================================

    return render(
        request,
        "admin_panel/edit_question.html",
        {
            "question": question,
            "next_url": next_url,
        }
    )
@login_required
@user_passes_test(is_admin)
def delete_selected_questions(request):

    if request.method == "POST":

        question_ids = request.POST.getlist("question_ids")

        if not question_ids:
            messages.warning(
                request,
                "Please select at least one question to delete."
            )

            return redirect("question_bank")

        questions = Question.objects.filter(
            id__in=question_ids
        )

        deleted_count = questions.count()

        questions.delete()

        messages.success(
            request,
            f"{deleted_count} question(s) deleted successfully."
        )

        return redirect("question_bank")

    return redirect("question_bank")
@login_required
@user_passes_test(is_admin_or_faculty)
def download_student_sample(request):

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"

    headers = [
        "Username",
        "Name",
        "Aadhaar Number",
    ]

    worksheet.append(headers)

    worksheet.append([
        "107225861001",
        "Student Name",
        "123456789012",
    ])

    worksheet.append([
        "107225861002",
        "Student Name 2",
        "234567890123",
    ])

    # Header formatting
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    # Column widths
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 20

    # Keep Username as text
    for row in worksheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=1
    ):
        for cell in row:
            cell.number_format = "@"

    # Keep Aadhaar as text
    for row in worksheet.iter_rows(
        min_row=2,
        min_col=3,
        max_col=3
    ):
        for cell in row:
            cell.number_format = "@"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_upload_sample.xlsx"'
    )

    workbook.save(response)

    return response
def production_check(request):

    User = get_user_model()

    username = "ravi"

    try:
        user = User.objects.filter(username=username).first()

        return JsonResponse({
            "database_engine": connection.vendor,
            "database_name": connection.settings_dict.get("NAME"),
            "user_exists": user is not None,
            "is_superuser": user.is_superuser if user else False,
            "is_staff": user.is_staff if user else False,
        })

    except Exception as e:
        return JsonResponse({
            "error": str(e)
        }, status=500)